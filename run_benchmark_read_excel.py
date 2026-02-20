import time
import glob
import os
import pandas as pd
from openpyxl import load_workbook

# pick a sample Excel file
files = glob.glob('Controle_Financeiro_*.xlsx')
if not files:
    print('NENHUM ARQUIVO Excel de amostra encontrado (Controle_Financeiro_*.xlsx)')
    raise SystemExit(1)

# choose the latest file by mtime
sample = max(files, key=os.path.getmtime)
print(f'Arquivo de amostra: {sample}')

# measure pandas read_excel
start = time.perf_counter()
df = pd.read_excel(sample, sheet_name='EXTRATO DETALHADO')
end = time.perf_counter()
print(f'pandas.read_excel: linhas={len(df)} cols={len(df.columns)} tempo={end-start:.4f}s')

# measure openpyxl load_workbook + counting rows
start = time.perf_counter()
wb = load_workbook(sample, read_only=True, data_only=True)
ws = wb['EXTRATO DETALHADO']
rows = sum(1 for _ in ws.iter_rows(values_only=True))
end = time.perf_counter()
print(f'openpyxl.load_workbook + iter_rows: linhas={rows} tempo={end-start:.4f}s')

# measure pandas read_excel with usecols minimal
start = time.perf_counter()
df_small = pd.read_excel(sample, sheet_name='EXTRATO DETALHADO', usecols=['Numero','Cliente','Bairro'])
end = time.perf_counter()
print(f'pandas.read_excel (usecols pequena): linhas={len(df_small)} cols={len(df_small.columns)} tempo={end-start:.4f}s')

print('\nBenchmark completo')
