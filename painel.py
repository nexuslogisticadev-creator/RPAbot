from collections import deque
import customtkinter as ctk
import tkinter as tk
import openpyxl
from tkinter import ttk, messagebox, simpledialog, filedialog
import json
import os
from datetime import datetime, timedelta
import subprocess
import threading
import queue
import sys
import unicodedata
import time
import shutil
import ctypes
from ctypes import wintypes

# ==================================================================================
#  SEÇÃO 1: IMPORTS E CONFIGURAÇÃO GLOBAL
# ==================================================================================
# Responsável por: Carregar todas as bibliotecas necessárias e definir
# constantes globais utilizadas pelo painel.
# ==================================================================================

try:
    import gspread
    from google.oauth2.service_account import Credentials
    from gspread.exceptions import WorksheetNotFound, APIError
    TEM_SHEETS = True
except ImportError:
    TEM_SHEETS = False


# --- BIBLIOTECAS DE JANELA ---
try:
    import win32gui
    import win32con
except ImportError:
    pass

# --- CALENDÁRIO ---
try:
    from tkcalendar import DateEntry
    TEM_CALENDARIO = True
except ImportError:
    TEM_CALENDARIO = False


# ================= DESIGN SYSTEM PRO =================
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

# Paleta de Cores Profissional
COR_BG_APP = "#0E0F11"           # Fundo principal mais escuro
COR_SIDEBAR = "#060708"          # Sidebar ultra escura
COR_CARD_BG = "#1C1D21"          # Cards com contraste melhor
COR_BORDA = "#2F3136"            # Bordas mais visíveis
COR_NEON = "#00E5FF"             # Azul neon para destaques
COR_ZE_AMARELO = "#FFD700"       # Amarelo Zé Delivery
COR_ZE_HOVER = "#FFC700"         # Hover mais vibrante
COR_SUCCESS = "#00E676"          # Verde de sucesso
COR_DANGER = "#FF1744"           # Vermelho de perigo
COR_TEXT_MAIN = "#FFFFFF"        # Texto principal branco puro
COR_TEXT_SEC = "#A0A0A0"         # Texto secundário mais claro
COR_TAB_HOVER = "#28292E"        # Hover de tabs

# Tipografia Profissional Aumentada
FONT_MAIN = ("Segoe UI", 14)           # Fonte principal maior
FONT_BOLD = ("Segoe UI Bold", 14)      # Negrito principal
FONT_TITLE = ("Segoe UI Bold", 26)     # Títulos maiores
FONT_MONO = ("Cascadia Code", 13)      # Fonte monospace moderna

ARQUIVO_COMANDO = 'comando_imprimir.txt'
ARQUIVO_CONFIG = 'config.json'
ARQUIVO_ESTOQUE = 'estoque.json'
ARQUIVO_ALERTAS = 'alertas_atraso.json'
ARQUIVO_FECHAMENTO_STATUS = 'fechamento_status.json'
ARQUIVO_MEMORIA_FECH = 'memoria_fechamento.json'

# ================= PERFORMANCE SETTINGS =================
# Ajuste fino para reduzir consumo quando o painel fica aberto.
AUTO_REFRESH_MS = 30000  # 30s - Auto-refresh do Excel
UI_QUEUE_INTERVAL_MS = 200  # Processamento de fila UI
UI_QUEUE_IDLE_MS = 500  # Economia de CPU quando idle
LOGS_REFRESH_ACTIVE_MS = 1000  # 1s - Logs só quando necessário
LOGS_REFRESH_IDLE_MS = 3000  # 3s - Economia máxima
ALERTAS_REFRESH_MS = 15000  # 15s - Verificação de alertas
MAX_ROWS_DISPLAY = 300  # Limite de linhas na tabela (performance)

# ================= FUNÇÕES AUXILIARES =================


def normalizar_texto(texto):
    """Normaliza texto removendo acentos, caracteres de combinação e espaços extras.

    Sempre retorna uma string em minúsculas. Não realiza leituras/escritas de arquivo.
    """
    try:
        s = str(texto)
        nfkd = unicodedata.normalize('NFKD', s)
        t = "".join([c for c in nfkd if not unicodedata.combining(c)])
        return t.lower().strip()
    except Exception:
        return str(texto).lower().strip()

def get_data_operacional():
    agora = datetime.now()
    if agora.hour < 10:
        agora -= timedelta(days=1)
    return agora.strftime("%d-%m-%Y")

def get_caminho_base():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

# ================= JANELA DE EDIÇÃO (MODAL) =================
class JanelaEdicao(ctk.CTkToplevel):
    def __init__(self, parent, dados_pedido, callback_salvar):
        super().__init__(parent)
        self.title(f"EDITAR PEDIDO #{dados_pedido['Numero']}")
        self.geometry("420x520")
        self.configure(fg_color=COR_BG_APP)
        self.callback = callback_salvar
        self.dados = dados_pedido
        self.transient(parent)
        self.grab_set()

        ctk.CTkLabel(self, text="EDITAR REGISTRO", font=FONT_TITLE,
                     text_color=COR_ZE_AMARELO).pack(pady=(20, 5))
        self.frm = ctk.CTkFrame(self, fg_color=COR_CARD_BG)
        self.frm.pack(padx=20, pady=10, fill="both", expand=True)

        self.criar_campo("Cliente:", dados_pedido['Cliente'], readonly=True)
        self.entry_bairro = self.criar_campo("Bairro:", dados_pedido['Bairro'])
        self.entry_valor = self.criar_campo(
            "Valor (R$):", str(dados_pedido['Valor']).replace("R$ ", ""))
        self.entry_motoboy = self.criar_campo(
            "Motoboy:", dados_pedido['Motoboy'])
        self.entry_status = self.criar_campo("Status:", dados_pedido['Status'])

        ctk.CTkButton(
            self, text="SALVAR", command=self.salvar,
            fg_color=COR_SUCCESS, text_color="#003300", height=50, font=FONT_BOLD
        ).pack(pady=20, padx=20, fill="x")

    def criar_campo(self, label, valor, readonly=False):
        f = ctk.CTkFrame(self.frm, fg_color="transparent")
        f.pack(fill="x", padx=15, pady=5)
        ctk.CTkLabel(f, text=label, text_color=COR_TEXT_SEC,
                     width=90, anchor="w", font=FONT_BOLD).pack(side="left")
        e = ctk.CTkEntry(f, fg_color="#111", border_color=COR_BORDA, height=35)
        e.insert(0, valor)
        if readonly:
            e.configure(state="disabled", fg_color="#222")
        e.pack(side="left", fill="x", expand=True)
        return e

    def salvar(self):
        novos = {
            'Bairro': self.entry_bairro.get(),
            'Valor (R$)': self.entry_valor.get().replace(",", "."),
            'Motoboy': self.entry_motoboy.get(),
            'Status': self.entry_status.get()
        }
        self.callback(self.dados['Numero'], novos)
        self.destroy()

# ================= INTERFACE PRINCIPAL =================
#
# ESTRUTURA DA CLASSE PainelUltra:
#
# SEÇÃO 1: INICIALIZAÇÃO E EVENTOS DO SISTEMA (linhas ~158-320)
#   __init__, _post_init_load, _auto_refresh_inteligente, _maximize_window,
#   _force_zoom_once, _on_unmap, _on_map_refresh, _on_resize, _apply_resize, _set_loading
#
# SEÇÃO 2: LAYOUT E INTERFACE (linhas ~329-620)
#   criar_menu_lateral, criar_botao_menu, criar_area_principal, _toggle_sidebar, mudar_aba,
#   criar_card_stat, criar_tabela_dark
#
# SEÇÃO 3: ABA FECHAMENTO (linhas ~622-1235)
#   setup_aba_fechamento, atualizar_dados_fechamento, _render_fechamento,
#   _carregar_dados_fechamento, _criar_linha_fechamento, calcular_fechamento_todos,
#   gerar_excel_fechamento, helpers (parse, calcular), motoboys
#
# SEÇÃO 4: ABA MONITOR/VALES (linhas ~1297-1600)
#   setup_aba_monitor, setup_aba_vales, carregar_tabela_vales, adicionar/editar/excluir_vale
#
# SEÇÃO 5: ABA ESTOQUE (linhas ~1584-1890)
#   setup_aba_estoque, carregar_estoque, add/del_produto, gerar_lista_compras
#
# SEÇÃO 6: ABA BI & CONFIGURAÇÃO (linhas ~1893-2040)
#   setup_aba_bi, atualizar_graficos_bi, gerar_mapa_calor, setup_aba_config
#
# SEÇÃO 7: LOGS, MOTOS E BAIRROS (linhas ~2042-2125)
#   setup_aba_logs, setup_aba_motos, setup_aba_bairros
#
# SEÇÃO 8: SISTEMA DE ROBÔ (linhas ~2157-2304)
#   buscar_robo_no_sistema, controlar_janela, toggle_robo, iniciar/parar_robo,
#   atualizar_logs_interface
#
# SEÇÃO 9: CACHE & ARQUIVO (linhas ~2309-2446)
#   carregar_config, salvar_config, atualizar_cache_bairros,
#   carregar_excel_cache, carregar_vales_cache, invalidar_cache_excel,
#   carregar_tabela, _coletar_dados_tabela, _render_tabela
#
# SEÇÃO 10: UTILITÁRIOS E FINAIS (linhas ~2644-2784)
#   mostrar_toast, atualizar_listas, salvar_motos, atualizar_bairros,
#   filtro_busca,  edição, impressão
#


class PainelUltra(ctk.CTk):
    def importar_estoque_do_excel(self, caminho_excel):
        """
        Importa o estoque do arquivo Excel de inventário e atualiza o estoque.json.
        Espera que o nome do produto esteja na coluna 'Produto' e o estoque conferido em 'Estoque Conferido'.
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(caminho_excel, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            col_produto = headers.index("Produto")
            col_estoque = headers.index("Estoque Conferido")
            novos_estoques = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                nome = str(row[col_produto]).strip() if row[col_produto] else None
                estoque = row[col_estoque]
                if nome and estoque is not None and str(estoque).replace('-', '').replace('.', '').isdigit():
                    try:
                        estoque_fisico = float(estoque)
                    except:
                        estoque_fisico = 0
                    novos_estoques.append({"nome": nome, "estoque_fisico": estoque_fisico})
            # Atualiza estoque.json
            with open("estoque.json", "w", encoding="utf-8") as f:
                json.dump(novos_estoques, f, ensure_ascii=False, indent=2)
            self.mostrar_toast(f"Estoque atualizado com {len(novos_estoques)} itens do Excel!", "success")
        except Exception as e:
            self.mostrar_toast(f"Erro ao importar estoque: {e}", "danger")

    def __init__(self):
        super().__init__()
        self.ESTOQUE_CATEGORIAS = [
            "🍺 CERVEJAS", "🍸 DESTILADOS", "🥤 NÃO ALCOÓLICOS", 
            "⚡ ENERGÉTICOS", "🍷 VINHOS & ICES", "🍟 MERCEARIA & DIVERSOS", "📦 OUTROS"
        ]
        self.title("ZÉ DELIVERY | CONTROL CENTER V8.0 PRO")
        self.configure(fg_color=COR_BG_APP)

        # ==================== CONFIGURAÇÃO PROFISSIONAL DE JANELA ====================
        screen_w = self.winfo_screenwidth()
        screen_h = self.winfo_screenheight()

        # Tenta usar a area util do Windows (evita barra de tarefas)
        work_area = None
        try:
            rect = wintypes.RECT()
            if ctypes.windll.user32.SystemParametersInfoW(0x0030, 0, ctypes.byref(rect), 0):
                work_area = (rect.left, rect.top, rect.right, rect.bottom)
        except Exception:
            work_area = None

        if work_area:
            left, top, right, bottom = work_area
            work_w = right - left
            work_h = bottom - top
            win_w = int(work_w * 0.94)
            win_h = int(work_h * 0.94)
            x = left + max(0, (work_w - win_w) // 2)
            y = top + max(0, (work_h - win_h) // 2)
            # Garante um offset mínimo do topo para evitar ficar muito alto
            min_offset = max(int(work_h * 0.06), 60)
            y = max(y, top + min_offset)
        else:
            # Fallback: usa tela inteira com margem segura
            safe_margin = 120
            win_w = int(screen_w * 0.92)
            win_h = int((screen_h - safe_margin) * 0.90)
            x = max(0, (screen_w - win_w) // 2)
            y = max(0, (screen_h - win_h - safe_margin) // 2)
            min_offset = max(int(screen_h * 0.03), 60)
            y = max(y, min_offset)

        self.geometry(f"{win_w}x{win_h}+{x}+{y}")
        
        # Tamanhos mínimos responsivos
        self.minsize(1000, 650)
        
        # Configuração de escala fixa (não muda dinamicamente)
        ctk.set_widget_scaling(1.0)
        
        # Estados e flags
        self._layout_mode = None
        self._sidebar_hidden = False
        self._defer_startup = True
        self._resize_after_id = None
        self._window_visible = True
        self._loading_monitor = False
        self._loading_fechamento = False
        self._loading_vales = False
        self._aba_atual = None

        self.processo_robo = None
        self.fila_logs = queue.Queue()
        self._ui_queue = queue.Queue()
        self.robo_rodando = False
        self.config_data = self.carregar_config()
        self.estoque_data = self.carregar_estoque()
        self.bairros_conhecidos = set()
        self.atualizar_cache_bairros()
        self.google_sheets_config = self._carregar_google_sheets_config()
        self.cache_detalhe_df = None
        self.cache_motos_df = None
        self.cache_excel_path = None
        self.cache_excel_mtime = None
        self.cache_vales_df = None
        self.cache_vales_path = None
        self.cache_vales_mtime = None
        # Rastreamento de mtime para cada aba (evitar recarregar mesmo arquivo)
        self.cache_fechamento_mtime = None
        self.cache_monitor_mtime = None
        self.search_after_id = None
        self.log_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "robo.log")
        self.log_tail_running = False
        self._log_buffer = deque(maxlen=2000)

        # Variáveis de Fechamento
        self.var_qtd8 = 0
        self.var_qtd11 = 0
        self.var_prod_total = 0.0
        self.var_garantia_calc = 0.0
        self.var_vale_total = 0.0
        self.var_modo_pagamento = tk.StringVar(value="Auto")
        self.data_var = tk.StringVar(value=get_data_operacional())
        self.memoria_fechamento = {}  # Cache para guardar horários digitados
        self._memoria_fech_all = {}
        self._carregar_memoria_fechamento(self.data_var.get())
        # Layout Principal: 0 = Sidebar, 1 = Conteúdo
        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # Deferir construção pesada da UI para agilizar o aparecimento da janela
        self.after(10, self._deferred_build)

    def _deferred_build(self):
        """Constrói o menu lateral, a área principal e inicia timers pesados.
        Executado via `after` para permitir que a janela apareça imediatamente.
        """
        try:
            self.criar_menu_lateral()
            self.criar_area_principal()

            self.frame_toast = ctk.CTkFrame(self, height=50, corner_radius=25, fg_color=COR_ZE_AMARELO)
            self.lbl_toast = ctk.CTkLabel(self.frame_toast, text="", font=FONT_BOLD, text_color="black")
            self.lbl_toast.pack(padx=30, pady=12)

            # Inicialização de processos (posposta para após a UI estar visível)
            self.after(100, self.atualizar_logs_interface)
            self.after(50, self._process_ui_queue)
            self.after(500, self.buscar_robo_no_sistema)
            # post init que seleciona aba inicial
            self.after(150, self._post_init_load)
            self.after(AUTO_REFRESH_MS, self._auto_refresh_inteligente)
        except Exception:
            pass

    # ═════════════════════════════════════════════════════════════════════════════
    # SEÇÃO 1: INICIALIZAÇÃO E EVENTOS DO SISTEMA
    # ═════════════════════════════════════════════════════════════════════════════

    # ==================================================================================
    #  SEÇÃO 2: MÉTODOS PRIVADOS E UTILITÁRIOS
    # ==================================================================================
    # Responsável por: Funções auxiliares, listeners de eventos, e métodos
    # internos que suportam a funcionalidade das abas.
    # ==================================================================================

    def _post_init_load(self):
        self._defer_startup = False
        self.mudar_aba("monitor")
        
        # Inicia timer para verificar alertas de atraso
        self._verificar_alertas_periodico()

    def _verificar_alertas_periodico(self):
        """Verifica alertas de atraso a cada 10 segundos se estiver na aba vales"""
        try:
            # Verifica qual aba está ativa
            current_tab = None
            for name, btn in self.botoes_menu.items():
                if btn.cget("border_width") == 1:
                    current_tab = name
                    break
            
            # Só recarrega se estiver na aba vales
            if current_tab == "vales":
                self.carregar_alertas_atraso()
        except Exception:
            pass
        
        # Reschedula usando constante configurável
        self.after(ALERTAS_REFRESH_MS, self._verificar_alertas_periodico)

    def _checar_fechamento_status(self):
        if not os.path.exists(ARQUIVO_FECHAMENTO_STATUS):
            return
        try:
            with open(ARQUIVO_FECHAMENTO_STATUS, 'r', encoding='utf-8') as f:
                data = json.load(f)
            status = str(data.get("status", "")).lower()
            if status == "ok":
                self.mostrar_toast("Fechamento enviado pelo robo.", "success")
            else:
                msg = data.get("msg") or "Falha ao enviar fechamento."
                self.mostrar_toast(str(msg), "error")
        except Exception:
            self.mostrar_toast("Falha ao ler status do fechamento.", "error")
        try:
            os.remove(ARQUIVO_FECHAMENTO_STATUS)
        except Exception:
            pass

    def _auto_refresh_inteligente(self):
        """
        Sistema de auto-refresh inteligente:
        - Verifica se o arquivo Excel mudou a cada 30s
        - Se mudou, recarrega APENAS a aba atual
        - Não recarrega se o arquivo não foi modificado (economiza CPU/recursos)
        - Pausa se a janela não está visível
        """
        try:
            self._checar_fechamento_status()

            # OTIMIZAÇÃO: Não atualiza se janela está minimizada/oculta
            if not getattr(self, '_window_visible', True):
                self.after(AUTO_REFRESH_MS, self._auto_refresh_inteligente)
                return
            
            arq = self._excel_path()
            if os.path.exists(arq):
                mtime = os.path.getmtime(arq)
                
                # Rastreia o mtime anterior para detectar mudanças
                if not hasattr(self, '_last_auto_refresh_mtime'):
                    self._last_auto_refresh_mtime = mtime
                elif mtime != self._last_auto_refresh_mtime:
                    # Arquivo mudou! Recarrega a aba atual
                    self._last_auto_refresh_mtime = mtime
                    current_tab = None
                    for name, btn in self.botoes_menu.items():
                        if btn.cget("border_width") == 1:
                            current_tab = name
                            break
                    
                    if current_tab == "monitor":
                        self.carregar_tabela()
                        self.carregar_alertas_atraso()  # Recarrega alertas também
                    elif current_tab == "fechamento":
                        self.atualizar_dados_fechamento()
                    elif current_tab == "vales":
                        self.carregar_tabela_vales()
        except Exception:
            pass  # Silencia erros de arquivo
        
        # Reschedula a próxima verificação
        self.after(AUTO_REFRESH_MS, self._auto_refresh_inteligente)

    def _on_unmap(self, _=None):
        """Janela minimizada/oculta - pausa atualizações para economizar recursos"""
        self._window_visible = False

    def _on_map_refresh(self, _=None):
        """Janela restaurada - retoma atualizações"""
        self._window_visible = True

    def _set_loading(self, active):
        if not hasattr(self, "barra_loading"):
            return
        if active:
            self.barra_loading.start()
        else:
            self.barra_loading.stop()
            self.barra_loading.set(0)

    def _format_alerta_auto_status(self):
        ativo = self.config_data.get("alerta_retirada_auto", False)
        return ("AUTO: ATIVO", COR_SUCCESS) if ativo else ("AUTO: DESLIGADO", COR_DANGER)

    def _atualizar_status_alerta_auto(self):
        if not hasattr(self, "lbl_alerta_auto_status"):
            return
        texto, cor = self._format_alerta_auto_status()
        self.lbl_alerta_auto_status.configure(text=texto, text_color=cor)

    # -------------------------------------------------------------------------
    # LAYOUT - SIDEBAR
    # -------------------------------------------------------------------------
    def criar_menu_lateral(self):
        self.sidebar = ctk.CTkFrame(self, width=280, corner_radius=0, fg_color=COR_SIDEBAR, border_width=0)
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_propagate(False)
        self.sidebar_body = self.sidebar

        # === RESPONSIVIDADE: LINHA 14 E A MOLA ===
        self.sidebar_body.grid_rowconfigure(14, weight=1)

        # Logo Profissional
        self.fr_logo = ctk.CTkFrame(self.sidebar_body, fg_color="transparent", height=80)
        self.fr_logo.grid(row=0, column=0, padx=28, pady=(28, 15), sticky="ew")
        ctk.CTkLabel(
            self.fr_logo, 
            text="ZÉ CONTROL", 
            font=("Segoe UI Black", 32), 
            text_color=COR_ZE_AMARELO
        ).pack(anchor="w")
        ctk.CTkLabel(
            self.fr_logo, 
            text="Enterprise Edition V8.0", 
            font=("Segoe UI", 11), 
            text_color=COR_TEXT_SEC
        ).pack(anchor="w", pady=(2, 0))

        # Separador visual
        ctk.CTkFrame(self.sidebar_body, height=1, fg_color=COR_BORDA).grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 15))

        # Botão Power com design melhorado
        self.btn_power = ctk.CTkButton(
            self.sidebar_body, 
            text="⚡ INICIAR SISTEMA", 
            command=self.toggle_robo,
            fg_color=COR_CARD_BG, 
            text_color=COR_SUCCESS, 
            font=("Segoe UI Bold", 15),
            height=55, 
            border_width=2, 
            border_color=COR_SUCCESS, 
            hover_color="#0d1f15",
            corner_radius=10
        )
        self.btn_power.grid(row=2, column=0, padx=20, pady=(0, 12), sticky="ew")

        # --- BOTÕES CHROME ---
        self.fr_chrome = ctk.CTkFrame(self.sidebar_body, fg_color="transparent")
        self.fr_chrome.grid(row=3, column=0, padx=20, pady=(0, 18), sticky="ew")

        self.btn_show_c = ctk.CTkButton(
            self.fr_chrome, text="👁️ MOSTRAR", width=110, height=32,
            command=lambda: self.controlar_janela("show"), fg_color="#1C1D21",
            border_width=1, border_color="#3A3B3F", font=("Segoe UI Semibold", 11), 
            hover_color="#28292E", corner_radius=8
        )
        self.btn_show_c.pack(side="left", padx=(0, 6), expand=True, fill="x")

        self.btn_hide_c = ctk.CTkButton(
            self.fr_chrome, text="👻 OCULTAR", width=110, height=32,
            command=lambda: self.controlar_janela("hide"), fg_color="#1C1D21",
            border_width=1, border_color="#3A3B3F", font=("Segoe UI Semibold", 11), 
            hover_color="#28292E", corner_radius=8
        )
        self.btn_hide_c.pack(side="right", padx=(6, 0), expand=True, fill="x")

        # Status com design melhorado
        self.fr_status = ctk.CTkFrame(self.sidebar_body, fg_color=COR_CARD_BG, corner_radius=8, height=50)
        self.fr_status.grid(row=4, column=0, pady=(0, 18), sticky="ew", padx=20)
        
        fr_status_content = ctk.CTkFrame(self.fr_status, fg_color="transparent")
        fr_status_content.pack(expand=True, pady=10, padx=15)
        
        self.lbl_status_dot = ctk.CTkLabel(
            fr_status_content, 
            text="●", 
            font=("Arial", 24), 
            text_color="#333"
        )
        self.lbl_status_dot.pack(side="left", padx=(0, 8))
        
        self.lbl_status_text = ctk.CTkLabel(
            fr_status_content, 
            text="SISTEMA OFFLINE", 
            text_color="#666", 
            font=("Segoe UI Bold", 13)
        )
        self.lbl_status_text.pack(side="left")

        # Separador + Label de Navegação
        ctk.CTkFrame(self.sidebar_body, height=1, fg_color=COR_BORDA).grid(row=5, column=0, sticky="ew", padx=20, pady=(0, 12))
        ctk.CTkLabel(
            self.sidebar_body, 
            text="NAVEGAÇÃO", 
            text_color=COR_TEXT_SEC, 
            font=("Segoe UI Semibold", 11)
        ).grid(row=6, column=0, pady=(0, 8), padx=28, sticky="w")

        self.botoes_menu = {}
        self.criar_botao_menu("📊  Dashboard", "monitor", 7)
        self.criar_botao_menu("💰  Fechamento", "fechamento", 8)
        self.criar_botao_menu("💸  Vales & Desc.", "vales", 9)
        self.criar_botao_menu("📦  Estoque", "estoque", 10)
        self.criar_botao_menu("🛵  Equipe", "motos", 11)
        self.criar_botao_menu("📍  Zonas", "bairros", 12)
        self.criar_botao_menu("🔑  PIX", "pix", 13)
        self.criar_botao_menu("💻  Terminal", "logs", 14)
        self.criar_botao_menu("⚙️  Configurações", "config", 15)

    def criar_botao_menu(self, texto, aba, row):
        btn = ctk.CTkButton(
            self.sidebar_body, text=texto, command=lambda: self.mudar_aba(aba),
            fg_color="transparent", text_color=COR_TEXT_SEC, hover_color=COR_TAB_HOVER,
            anchor="w", font=FONT_MAIN, height=40, corner_radius=6
        )
        btn.grid(row=row, column=0, padx=15, pady=1, sticky="ew")
        self.botoes_menu[aba] = btn

    # -------------------------------------------------------------------------
    # LAYOUT - ÁREA PRINCIPAL
    # -------------------------------------------------------------------------

    # ==================================================================================
    #  SEÇÃO 3: LAYOUT PRINCIPAL (MENU + ÁREA PRINCIPAL)
    # ==================================================================================
    # Responsável por: Criação da estrutura visual principal (menu lateral,
    # abas, e estrutura de navegação do painel).
    # ==================================================================================

    def criar_area_principal(self):
        self.main_container = ctk.CTkFrame(self, fg_color=COR_BG_APP)
        self.main_container.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
        self.main_container.grid_rowconfigure(1, weight=1)
        self.main_container.grid_columnconfigure(0, weight=1)
        

        self.barra_loading = ctk.CTkProgressBar(self.main_container, height=3, progress_color=COR_ZE_AMARELO)
        self.barra_loading.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        self.barra_loading.set(0)

        self.btn_sidebar_toggle = ctk.CTkButton(
            self.main_container,
            text="☰",
            width=36,
            height=32,
            fg_color=COR_CARD_BG,
            text_color=COR_TEXT_MAIN,
            command=self._toggle_sidebar
        )
        self.btn_sidebar_toggle.place_forget()

        self.frames = {}
        # Cria apenas os contêineres das abas; conteúdo será carregado on-demand
        tab_names = ["monitor", "vales", "fechamento", "estoque", "logs", "motos", "pix", "bairros", "config"]
        for nome in tab_names:
            fr = ctk.CTkFrame(self.main_container, fg_color=COR_BG_APP)
            self.frames[nome] = fr
            fr.grid(row=1, column=0, sticky="nsew")

        # Mapeia funções de setup para carregamento preguiçoso
        self._tab_setup_funcs = {
            "vales": self.setup_aba_vales,
            "monitor": self.setup_aba_monitor,
            "fechamento": self.setup_aba_fechamento,
            "estoque": self.setup_aba_estoque,
            "logs": self.setup_aba_logs,
            "motos": self.setup_aba_motos,
            "pix": self.setup_aba_pix,
            "bairros": self.setup_aba_bairros,
            "config": self.setup_aba_config,
        }
        self._tab_setup_done = {name: False for name in tab_names}
        # Aba manual removida

    def _on_resize(self, event):
        """Sistema de resize responsivo sem loops - Ajusta layout baseado na largura"""
        if event.widget is not self:
            return
        
        # Debounce: Cancela resize anterior se ainda não executou
        if self._resize_after_id:
            try:
                self.after_cancel(self._resize_after_id)
            except Exception:
                pass
        
        # Aguarda 400ms de inatividade antes de aplicar mudanças
        self._resize_after_id = self.after(400, lambda w=event.width: self._apply_resize(w))

    def _apply_resize(self, w):
        """Aplica ajustes responsivos baseado na largura da janela"""
        self._resize_after_id = None
        h = self.winfo_height()
        compact_h = h < 720
        
        # Define breakpoints profissionais
        if w < 1000:
            mode = "compact"
            sidebar_w = 220
            pad = 12
        elif w < 1400:
            mode = "normal"
            sidebar_w = 260
            pad = 18
        else:
            mode = "wide"
            sidebar_w = 280
            pad = 22

        # Evita re-aplicar o mesmo modo (previne loops)
        if mode == self._layout_mode:
            return
        self._layout_mode = mode

        # Ajusta largura da sidebar e padding
        self.sidebar.configure(width=sidebar_w)
        if hasattr(self, "sidebar_scroll"):
            self.sidebar_scroll.configure(width=sidebar_w)
        self.main_container.grid_configure(padx=pad, pady=pad)

        # Em modo compacto, esconde sidebar e mostra botão toggle
        if mode == "compact":
            if not self._sidebar_hidden:
                self.sidebar.grid_remove()
                self._sidebar_hidden = True
            self.btn_sidebar_toggle.place(x=8, y=8)
        else:
            if self._sidebar_hidden:
                self.sidebar.grid()
                self._sidebar_hidden = False
            self.btn_sidebar_toggle.place_forget()

        # Ajusta widgets específicos se existirem
        if hasattr(self, "ent_busca"):
            ent_busca_w = 150 if mode == "compact" else (180 if mode == "normal" else 220)
            self.ent_busca.configure(width=ent_busca_w)
        
        if hasattr(self, "combo_motos_ativos"):
            combo_w = 160 if mode == "compact" else (190 if mode == "normal" else 220)
            self.combo_motos_ativos.configure(width=combo_w)
        
        if hasattr(self, "ent_prod"):
            prod_w = 200 if mode == "compact" else (230 if mode == "normal" else 260)
            self.ent_prod.configure(width=prod_w)
        
        if hasattr(self, "ent_qtd"):
            qtd_w = 75 if mode == "compact" else (85 if mode == "normal" else 95)
            self.ent_qtd.configure(width=qtd_w)
        
        if hasattr(self, "combo_vale_moto"):
            vale_w = 180 if mode == "compact" else (210 if mode == "normal" else 240)
            try:
                self.combo_vale_moto.configure(width=vale_w)
            except Exception:
                pass
        if hasattr(self, "combo_moto_vales"):
            combo_vales_w = 200 if mode == "compact" else (220 if mode == "normal" else 240)
            self.combo_moto_vales.configure(width=combo_vales_w)

        # Ajustes verticais para telas baixas
        if hasattr(self, "btn_power"):
            self.btn_power.configure(height=48 if compact_h else 55)
            self.btn_power.configure(font=("Segoe UI Bold", 14 if compact_h else 15))
        if hasattr(self, "fr_logo"):
            self.fr_logo.grid_configure(pady=(20, 10) if compact_h else (28, 15))
        if hasattr(self, "fr_chrome"):
            self.fr_chrome.grid_configure(pady=(0, 10) if compact_h else (0, 18))
        if hasattr(self, "btn_show_c"):
            self.btn_show_c.configure(height=28 if compact_h else 32)
        if hasattr(self, "btn_hide_c"):
            self.btn_hide_c.configure(height=28 if compact_h else 32)
        if hasattr(self, "fr_status"):
            self.fr_status.configure(height=44 if compact_h else 50)
        if hasattr(self, "lbl_status_text"):
            self.lbl_status_text.configure(font=("Segoe UI Bold", 12 if compact_h else 13))
        if hasattr(self, "fr_data_ops"):
            self.fr_data_ops.pack_configure(padx=(6, 0) if compact_h else (10, 0))
        if hasattr(self, "btn_atualizar_data"):
            self.btn_atualizar_data.configure(height=32 if compact_h else 36)

        if hasattr(self, "tree_retirada"):
            if mode == "compact":
                widths = [70, 80, 160, 120, 90, 110, 80]
            elif mode == "normal":
                widths = [80, 90, 200, 140, 100, 130, 90]
            else:
                widths = [90, 100, 240, 160, 110, 150, 100]
            cols = ["Hora", "Numero", "Cliente", "Bairro", "Status", "Motoboy", "Valor"]
            for tree in [self.tree_retirada, self.tree_cancelado, self.tree_entrega]:
                for col, w in zip(cols, widths):
                    tree.column(col, width=w)

        if hasattr(self, "tree_vales"):
            if mode == "compact":
                widths = [0, 90, 180, 120, 200]
            elif mode == "normal":
                widths = [0, 100, 220, 140, 240]
            else:
                widths = [0, 110, 250, 150, 300]
            cols = ["ID", "HORA", "MOTOBOY", "VALOR", "MOTIVO"]
            for col, w in zip(cols, widths):
                self.tree_vales.column(col, width=w)

        if hasattr(self, "tree_estoque"):
            if mode == "compact":
                widths = [220, 90, 110, 90, 130]
            elif mode == "normal":
                widths = [260, 100, 120, 100, 150]
            else:
                widths = [300, 110, 140, 110, 170]
            cols = ["PRODUTO", "NIVEL", "STATUS", "PRECO", "FORNECEDOR"]
            for col, w in zip(cols, widths):
                self.tree_estoque.column(col, width=w)

        if hasattr(self, "fr_vales_top") and hasattr(self, "fr_vales_left") and hasattr(self, "fr_vales_actions"):
            if mode == "compact":
                self.fr_vales_left.grid_configure(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))
                self.fr_vales_actions.grid_configure(row=1, column=0, columnspan=2, sticky="ew")
            else:
                self.fr_vales_left.grid_configure(row=0, column=0, columnspan=1, sticky="w", pady=0)
                self.fr_vales_actions.grid_configure(row=0, column=1, columnspan=1, sticky="e")

    def _toggle_sidebar(self):
        if self._sidebar_hidden:
            self.sidebar.grid()
            self._sidebar_hidden = False
        else:
            self.sidebar.grid_remove()
            self._sidebar_hidden = True

    def mudar_aba(self, nome_aba):
        # Garante que a aba tenha seu conteúdo inicializado antes de exibir
        try:
            if not getattr(self, '_tab_setup_done', {}).get(nome_aba, False):
                setup_fn = getattr(self, '_tab_setup_funcs', {}).get(nome_aba)
                if setup_fn:
                    setup_fn(self.frames[nome_aba])
                    self._tab_setup_done[nome_aba] = True
        except Exception:
            pass

        self.frames[nome_aba].tkraise()
        self._aba_atual = nome_aba
        for nome, btn in self.botoes_menu.items():
            if nome == nome_aba:
                btn.configure(fg_color=COR_CARD_BG, text_color="white", border_width=1, border_color=COR_NEON)
            else:
                btn.configure(fg_color="transparent", text_color=COR_TEXT_SEC, border_width=0)

        if nome_aba == "fechamento": self.atualizar_dados_fechamento()
        if nome_aba == "monitor":
            self.carregar_tabela()
        if nome_aba == "logs": self.buscar_robo_no_sistema()
        if nome_aba == "pix": self.atualizar_lista_pix()
        if nome_aba == "vales":
            self.atualizar_lista_motoboys_vales()
            self.carregar_tabela_vales()
            self.carregar_alertas_atraso()  # Carrega alertas ao abrir vales
    
    # ═════════════════════════════════════════════════════════════════════════════
    # SEÇÃO 3: ABA FECHAMENTO
    # ═════════════════════════════════════════════════════════════════════════════
    # Cálculo automático de produção vs garantia para pagamento de motoboys
    # ─────────────────────────────────────────────────────────────────────────────
    # ==================================================================================
    #  SEÇÃO 5: ABA FECHAMENTO & PAGAMENTO
    # ==================================================================================
    # Responsável por: Cálculo de fechamento diário, processamento de pagamentos
    # a motoboys, tratamento de vales e descontos.
    # ==================================================================================

    def setup_aba_fechamento(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(2, weight=1)

        ctk.CTkLabel(
            parent,
            text="FECHAMENTO AUTOMATICO (PRODUCAO X GARANTIA)",
            font=FONT_TITLE,
            text_color=COR_ZE_AMARELO
        ).grid(row=0, column=0, pady=(20, 10))

        fr_acoes = ctk.CTkFrame(parent, fg_color=COR_CARD_BG, height=60)
        fr_acoes.grid(row=1, column=0, sticky="ew", padx=20, pady=10)
        fr_acoes.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(fr_acoes, text="Dica: preencha apenas os garantidos.", font=FONT_MAIN, text_color=COR_TEXT_SEC).pack(side="left", padx=15)

        ctk.CTkButton(
            fr_acoes,
            text="CALCULAR FINAL",
            command=self.calcular_fechamento_todos,
            height=40,
            fg_color=COR_ZE_AMARELO,
            text_color="black",
            font=FONT_BOLD
        ).pack(side="right", padx=10)

        ctk.CTkButton(
            fr_acoes,
            text="GERAR EXCEL",
            command=self.gerar_excel_fechamento,
            height=40,
            fg_color="#333",
            text_color="white",
            font=FONT_BOLD
        ).pack(side="right", padx=10)

        ctk.CTkButton(
            fr_acoes,
            text="ENVIAR FECHAMENTO",
            command=self.enviar_fechamento_manual,
            height=40,
            fg_color=COR_SUCCESS,
            text_color="#003300",
            font=FONT_BOLD
        ).pack(side="right", padx=10)

        self.fr_fechamento_table = ctk.CTkScrollableFrame(parent, fg_color=COR_BG_APP)
        self.fr_fechamento_table.grid(row=2, column=0, sticky="nsew", padx=20, pady=10)
        self.fr_fechamento_table.grid_columnconfigure(0, weight=1)

        self.fechamento_rows = []
        self.fechamento_rows_by_name = {}
        if not self._defer_startup:
            self.atualizar_dados_fechamento()
    def atualizar_dados_fechamento(self):
        if self._loading_fechamento:
            return
        
        arq = self._excel_path()
        
        # Verificação inteligente: só carrega se mudou
        try:
            mtime = os.path.getmtime(arq)
        except OSError:
            return
        
        # Se arquivo não mudou, não recarrega
        if mtime == self.cache_fechamento_mtime:
            return
        
        self._loading_fechamento = True
        self._set_loading(True)
        self.cache_fechamento_mtime = mtime

        # carregamos a memória de horários para a data selecionada antes de renderizar
        try:
            self._carregar_memoria_fechamento(self.data_var.get())
        except Exception:
            pass

        def worker():
            dados = self._carregar_dados_fechamento()
            self._enqueue_ui(lambda d=dados: self._render_fechamento(d))

        threading.Thread(target=worker, daemon=True).start()

    def _render_fechamento(self, dados):
        self._loading_fechamento = False
        self._set_loading(False)
        self.fechamento_rows = []
        self.fechamento_rows_by_name = {}
        self._limpar_fechamento_tabela()
        self._montar_cabecalho_fechamento()

        if not dados:
            self.mostrar_toast("Nenhum motoboy encontrado para esta data.", "info")
            return

        names = list(dados.keys())
        total = len(names)

        # Se muitos motoboys, renderiza apenas os primeiros para evitar travamento
        if total > MAX_ROWS_DISPLAY:
            to_show = names[:MAX_ROWS_DISPLAY]
            remaining = names[MAX_ROWS_DISPLAY:]
            for nome in to_show:
                self._criar_linha_fechamento(nome, dados[nome])

            # guarda pendentes para carregamento sob demanda
            self._fechamento_pending = {n: dados[n] for n in remaining}

            # botão para carregar o restante (carrega em batches via _enqueue_ui)
            fr_btn = ctk.CTkFrame(self.fr_fechamento_table, fg_color="transparent")
            fr_btn.grid(row=MAX_ROWS_DISPLAY + 1, column=0, sticky="ew", padx=6, pady=10)
            btn = ctk.CTkButton(fr_btn, text=f"Carregar mais {len(remaining)} motoboys",
                                command=self._carregar_mais_fechamento,
                                fg_color=COR_NEON, text_color="black", height=38)
            btn.pack()
            self._btn_carregar_mais_fechamento = btn
        else:
            for nome in names:
                self._criar_linha_fechamento(nome, dados[nome])

        self.calcular_fechamento_todos()

    def _limpar_fechamento_tabela(self):
        if not hasattr(self, "fr_fechamento_table"):
            return
        for child in self.fr_fechamento_table.winfo_children():
            child.destroy()

    def _montar_cabecalho_fechamento(self):
        cab = ctk.CTkFrame(self.fr_fechamento_table, fg_color=COR_CARD_BG, border_width=1, border_color=COR_BORDA)
        cab.grid(row=0, column=0, sticky="ew", padx=2, pady=(2, 6))
        cab.grid_columnconfigure((0, 1, 2, 3, 4, 5, 6, 7, 8), weight=1)

        labels = [
            "MOTOBOY", "QTD 8", "QTD 11", "TOTAL PROD", "HORA INI",
            "HORA FIM", "VALOR GAR", "A PAGAR", "PIX"
        ]
        for idx, texto in enumerate(labels):
            ctk.CTkLabel(cab, text=texto, font=FONT_BOLD, text_color=COR_TEXT_SEC).grid(
                row=0, column=idx, padx=6, pady=8, sticky="ew"
            )

    def _carregar_dados_fechamento(self):
        arq = self._excel_path()
        if not os.path.exists(arq):
            return {}

        try:
            df, df_m = self.carregar_excel_cache(arq)
        except Exception:
            return {}

        if df is None:
            return {}

        dados = {}

        for _, row in df.iterrows():
            try:
                status = str(row['Status']).upper()
                if "CANCEL" in status or "ABANDONED" in status:
                    continue
                nome = str(row['Motoboy']).strip()
                if not nome or nome.upper() == "RETIRADA":
                    continue

                val = float(row['Valor (R$)']) if row['Valor (R$)'] else 0.0
                hora = str(row['Hora']).strip()

                if nome not in dados:
                    dados[nome] = {
                        "qtd8": 0,
                        "qtd11": 0,
                        "prod_total": 0.0,
                        "entregas": []
                    }

                if abs(val - 8.0) < 0.1:
                    dados[nome]["qtd8"] += 1
                elif abs(val - 11.0) < 0.1:
                    dados[nome]["qtd11"] += 1

                dados[nome]["prod_total"] += val
                dados[nome]["entregas"].append({"hora": hora, "valor": val})
            except Exception:
                continue

        if df_m is not None and "MOTOBOY" in df_m.columns:
            for _, row in df_m.iterrows():
                nome = str(row.get("MOTOBOY", "")).strip()
                if nome and nome.upper() != "RETIRADA" and nome not in dados:
                    dados[nome] = {
                        "qtd8": int(row.get("QTD R$ 8,00", 0) or 0),
                        "qtd11": int(row.get("QTD R$ 11,00", 0) or 0),
                        "prod_total": float(row.get("TOTAL A PAGAR (R$)", 0.0) or 0.0),
                        "entregas": []
                    }

        return dict(sorted(dados.items()))

    # ==================== MEMÓRIA DE FECHAMENTO (PERSISTÊNCIA POR DATA) ====================
    def _carregar_memoria_fechamento(self, data_operacional=None):
        try:
            if data_operacional is None:
                data_operacional = get_data_operacional()
            if not os.path.exists(ARQUIVO_MEMORIA_FECH):
                self._memoria_fech_all = {}
                self.memoria_fechamento = {}
                return
            with open(ARQUIVO_MEMORIA_FECH, 'r', encoding='utf-8') as f:
                self._memoria_fech_all = json.load(f)
        except Exception:
            self._memoria_fech_all = {}

        # carrega apenas a data solicitada
        self.memoria_fechamento = self._memoria_fech_all.get(data_operacional, {}) or {}

    def _salvar_memoria_fechamento(self, data_operacional=None):
        try:
            if data_operacional is None:
                data_operacional = get_data_operacional()
            # garante que tenhamos o dicionário completo atualizado
            try:
                if os.path.exists(ARQUIVO_MEMORIA_FECH):
                    with open(ARQUIVO_MEMORIA_FECH, 'r', encoding='utf-8') as f:
                        self._memoria_fech_all = json.load(f)
                else:
                    self._memoria_fech_all = {}
            except Exception:
                self._memoria_fech_all = {}

            self._memoria_fech_all[data_operacional] = self.memoria_fechamento or {}
            with open(ARQUIVO_MEMORIA_FECH, 'w', encoding='utf-8') as f:
                json.dump(self._memoria_fech_all, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"❌ Erro salvando memória de fechamento: {e}")
        

    def _criar_linha_fechamento(self, nome, info):
        linha_idx = len(self.fechamento_rows) + 1
        fr = ctk.CTkFrame(self.fr_fechamento_table, fg_color=COR_CARD_BG, border_width=1, border_color=COR_BORDA)
        fr.grid(row=linha_idx, column=0, sticky="ew", padx=2, pady=2)
        fr.grid_columnconfigure((0, 1, 2, 3, 4, 5, 6, 7, 8), weight=1)

        lbl_nome = ctk.CTkLabel(fr, text=nome, font=FONT_BOLD, text_color=COR_TEXT_MAIN)
        lbl_nome.grid(row=0, column=0, padx=6, pady=6, sticky="w")

        lbl_q8 = ctk.CTkLabel(fr, text=str(info["qtd8"]), font=FONT_MAIN)
        lbl_q8.grid(row=0, column=1, padx=6, pady=6, sticky="ew")

        lbl_q11 = ctk.CTkLabel(fr, text=str(info["qtd11"]), font=FONT_MAIN)
        lbl_q11.grid(row=0, column=2, padx=6, pady=6, sticky="ew")

        lbl_total = ctk.CTkLabel(fr, text=f"R$ {info['prod_total']:.2f}", font=FONT_MAIN)
        lbl_total.grid(row=0, column=3, padx=6, pady=6, sticky="ew")

        ent_ini = ctk.CTkEntry(fr, width=90, placeholder_text="HH:MM")
        ent_ini.grid(row=0, column=4, padx=6, pady=6, sticky="ew")

        ent_fim = ctk.CTkEntry(fr, width=90, placeholder_text="HH:MM")
        ent_fim.grid(row=0, column=5, padx=6, pady=6, sticky="ew")

        ent_gar = ctk.CTkEntry(fr, width=100, placeholder_text="Ex: 120.00")
        ent_gar.grid(row=0, column=6, padx=6, pady=6, sticky="ew")

        # Restaura valores da memória se existirem
        if nome in self.memoria_fechamento:
            mem = self.memoria_fechamento[nome]
            if mem.get("ini"): ent_ini.insert(0, mem["ini"])
            if mem.get("fim"): ent_fim.insert(0, mem["fim"])
            if mem.get("gar"): ent_gar.insert(0, mem["gar"])

        lbl_pagar = ctk.CTkLabel(fr, text="R$ 0.00", font=FONT_BOLD, text_color=COR_SUCCESS)
        lbl_pagar.grid(row=0, column=7, padx=6, pady=6, sticky="ew")

        btn_pix = ctk.CTkButton(
            fr,
            text="PIX",
            width=60,
            command=lambda n=nome: self._copiar_pix_motoboy(n),
            fg_color="#333",
            text_color="white",
            font=FONT_BOLD
        )
        btn_pix.grid(row=0, column=8, padx=6, pady=6, sticky="ew")

        ent_ini.bind("<KeyRelease>", lambda _, n=nome: self._recalcular_fechamento_linha(n, False))
        ent_fim.bind("<KeyRelease>", lambda _, n=nome: self._recalcular_fechamento_linha(n, False))
        ent_gar.bind("<KeyRelease>", lambda _, n=nome: self._recalcular_fechamento_linha(n, False))

        row = {
            "nome": nome,
            "qtd8": info["qtd8"],
            "qtd11": info["qtd11"],
            "prod_total": info["prod_total"],
            "entregas": info["entregas"],
            "ent_inicio": ent_ini,
            "ent_fim": ent_fim,
            "ent_garantia": ent_gar,
            "lbl_pagar": lbl_pagar,
            "calc": {}
        }

        self.fechamento_rows.append(row)
        self.fechamento_rows_by_name[nome] = row

    def _parse_float(self, texto):
        if texto is None:
            return None
        try:
            return float(str(texto).replace(",", ".").strip())
        except Exception:
            return None

    def _parse_hora(self, texto):
        try:
            return datetime.strptime(texto, "%H:%M")
        except Exception:
            return None

    def _calcular_garantia_valor(self, t_in, t_out):
        if not t_in or not t_out:
            return 0.0
        if t_out <= t_in:
            return 0.0
        if t_in.strftime("%H:%M") == "10:00" and t_out.strftime("%H:%M") == "19:00":
            return 120.0
        horas = (t_out - t_in).total_seconds() / 3600
        return max(horas * 15.0, 0.0)

    def _recalcular_fechamento_linha(self, nome, mostrar_erros):
        row = self.fechamento_rows_by_name.get(nome)
        if not row:
            return False

        hora_ini_txt = row["ent_inicio"].get().strip()
        hora_fim_txt = row["ent_fim"].get().strip()
        gar_txt = row["ent_garantia"].get().strip()

        # Salva na memória para restaurar se a UI recarregar
        if nome not in self.memoria_fechamento:
            self.memoria_fechamento[nome] = {}
        self.memoria_fechamento[nome]['ini'] = hora_ini_txt
        self.memoria_fechamento[nome]['fim'] = hora_fim_txt
        self.memoria_fechamento[nome]['gar'] = gar_txt
        try:
            self._salvar_memoria_fechamento(self.data_var.get())
        except Exception:
            pass

        if not hora_ini_txt or not hora_fim_txt:
            total = row["prod_total"]
            row["lbl_pagar"].configure(text=f"R$ {total:.2f}")
            row["calc"] = {
                "qtd8_fora": row["qtd8"],
                "qtd11_fora": row["qtd11"],
                "qtd8_dentro": 0,
                "qtd11_dentro": 0,
                "prod_fora": row["prod_total"],
                "prod_dentro": 0.0,
                "garantia": 0.0,
                "usar_garantia": False,
                "total": total
            }
            return True

        t_in = self._parse_hora(hora_ini_txt)
        t_out = self._parse_hora(hora_fim_txt)
        if not t_in or not t_out or t_out <= t_in:
            if mostrar_erros:
                self.mostrar_toast(f"Horario invalido em {nome} (use HH:MM).", "error")
            total = row["prod_total"]
            row["lbl_pagar"].configure(text=f"R$ {total:.2f}")
            row["calc"] = {
                "qtd8_fora": row["qtd8"],
                "qtd11_fora": row["qtd11"],
                "qtd8_dentro": 0,
                "qtd11_dentro": 0,
                "prod_fora": row["prod_total"],
                "prod_dentro": 0.0,
                "garantia": 0.0,
                "usar_garantia": False,
                "total": total
            }
            return False

        qtd8_dentro = 0
        qtd11_dentro = 0
        qtd8_fora = 0
        qtd11_fora = 0
        prod_dentro = 0.0
        prod_fora = 0.0

        for entrega in row["entregas"]:
            t_ped = self._parse_hora(str(entrega.get("hora", "")))
            val = float(entrega.get("valor", 0.0) or 0.0)

            # Regra: conta como dentro se a hora do pedido for menor ou igual ao fim do garantido
            dentro = t_ped is not None and t_ped <= t_out
            if dentro:
                prod_dentro += val
                if abs(val - 8.0) < 0.1:
                    qtd8_dentro += 1
                elif abs(val - 11.0) < 0.1:
                    qtd11_dentro += 1
            else:
                prod_fora += val
                if abs(val - 8.0) < 0.1:
                    qtd8_fora += 1
                elif abs(val - 11.0) < 0.1:
                    qtd11_fora += 1

        gar_valor = self._parse_float(gar_txt)
        if gar_valor is None:
            gar_valor = self._calcular_garantia_valor(t_in, t_out)

        usar_garantia = gar_valor > prod_dentro
        if usar_garantia:
            total = gar_valor + prod_fora
        else:
            total = prod_dentro + prod_fora

        row["lbl_pagar"].configure(text=f"R$ {total:.2f}")
        row["calc"] = {
            "qtd8_fora": qtd8_fora,
            "qtd11_fora": qtd11_fora,
            "qtd8_dentro": qtd8_dentro,
            "qtd11_dentro": qtd11_dentro,
            "prod_fora": prod_fora,
            "prod_dentro": prod_dentro,
            "garantia": gar_valor,
            "usar_garantia": usar_garantia,
            "total": total
        }
        return True

    def calcular_fechamento_todos(self):
        if not self.fechamento_rows:
            return
        ok = True
        for row in self.fechamento_rows:
            if not self._recalcular_fechamento_linha(row["nome"], True):
                ok = False
        if ok:
            self.mostrar_toast("Fechamento atualizado.", "success")

    def _carregar_mais_fechamento(self):
        """Enfileira a criação das linhas restantes para não travar a UI."""
        pending = getattr(self, "_fechamento_pending", {})
        if not pending:
            return

        # Enfileira cada criação; _process_ui_queue irá processar em batches
        for nome, info in list(pending.items()):
            self._enqueue_ui(lambda n=nome, i=info: self._criar_linha_fechamento(n, i))
            try:
                del self._fechamento_pending[nome]
            except Exception:
                pass

        # remove botão de carregar mais (será recriado em caso de mais pendentes)
        if hasattr(self, "_btn_carregar_mais_fechamento"):
            try:
                self._btn_carregar_mais_fechamento.destroy()
            except Exception:
                pass
            delattr(self, "_btn_carregar_mais_fechamento")

    def _obter_pix_motoboy(self, nome):
        pix_map = self.config_data.get("pix_motoboys", {})
        if not nome:
            return ""
        
        # Limpa o nome de entrada
        nome_limpo = str(nome).strip()
        
        # Tenta busca direta primeiro
        pix = pix_map.get(nome_limpo)
        if pix:
            return str(pix).strip()
        
        # Tenta busca normalizada (sem acentos, minúsculas)
        nome_norm = normalizar_texto(nome_limpo)
        for k, v in pix_map.items():
            k_norm = normalizar_texto(k)
            if k_norm == nome_norm:
                return str(v).strip()
        
        return ""

    def _copiar_pix_motoboy(self, nome):
        pix = self._obter_pix_motoboy(nome)
        if not pix:
            pix = simpledialog.askstring("PIX", f"Digite a chave PIX de {nome}:")
            if not pix:
                return
            self.config_data.setdefault("pix_motoboys", {})[nome] = pix.strip()
            self.salvar_config()
        self.clipboard_clear()
        self.clipboard_append(pix)
        self.mostrar_toast("PIX copiado!", "success")

    def _carregar_google_sheets_config(self):
        """Carrega configuracoes do Google Sheets do config.json com valores padrão."""
        config = self.config_data.get("google_sheets", {})
        defaults = {
            "service_account_json": r"C:\Users\Usuario\Desktop\teste_novo\gen-lang-client-0592009269-3d0b6d104f80.json",
            "sheets_id": "1f716jdjISk1Xlu6wj5weAiFRlt_b_IDPixS1GuSfudE"
        }
        return {**defaults, **config}

    def gerar_excel_fechamento(self):
        if not self.fechamento_rows:
            self.mostrar_toast("Nao ha dados para exportar.", "error")
            return
        if not TEM_SHEETS:
            self.mostrar_toast("Instale gspread e google-auth para usar o Sheets.", "error")
            return

        cred_path = self.google_sheets_config.get("service_account_json", "")
        if not cred_path or not os.path.exists(cred_path):
            self.mostrar_toast("Arquivo JSON da conta de servico nao encontrado.", "error")
            return

        sheet_id = self.google_sheets_config.get("sheets_id", "")
        if not sheet_id:
            self.mostrar_toast("ID do Google Sheets nao configurado.", "error")
            return
        sheet_title = self._obter_nome_aba_sheets()
        if not sheet_title:
            self.mostrar_toast("Data invalida para nome da aba.", "error")
            return

        try:
            scopes = [
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive"
            ]
            creds = Credentials.from_service_account_file(cred_path, scopes=scopes)
            client = gspread.authorize(creds)
            sh = client.open_by_key(sheet_id)

            try:
                ws = sh.worksheet(sheet_title)
                ws.clear()
            except WorksheetNotFound:
                ws = sh.add_worksheet(title=sheet_title, rows=200, cols=14)
            except APIError:
                sheet_title = sheet_title.replace("/", "-")
                try:
                    ws = sh.worksheet(sheet_title)
                    ws.clear()
                except WorksheetNotFound:
                    ws = sh.add_worksheet(title=sheet_title, rows=200, cols=14)

            cabecalhos = [
                "Motoboy", "8", "11", "Entregas", "Pago",
                "PIX", "TOTAL", "Val", "Garantido",
                "8", "11", "", "Inicio", "Fim"
            ]

            linhas = [cabecalhos]
            for row in self.fechamento_rows:
                nome = row["nome"]
                calc = row.get("calc", {})
                vale = self.calcular_total_vales_moto(nome)
                pix = self._obter_pix_motoboy(nome)

                hora_ini = row["ent_inicio"].get().strip()
                hora_fim = row["ent_fim"].get().strip()

                if calc.get("usar_garantia"):
                    qtd8_out = calc.get("qtd8_fora", 0)
                    qtd11_out = calc.get("qtd11_fora", 0)
                    qtd8_in = calc.get("qtd8_dentro", 0)
                    qtd11_in = calc.get("qtd11_dentro", 0)
                    garantia = calc.get("garantia", 0.0)
                else:
                    qtd8_out = row["qtd8"]
                    qtd11_out = row["qtd11"]
                    qtd8_in = 0
                    qtd11_in = 0
                    garantia = 0.0

                total_entregas = qtd8_out + qtd11_out + qtd8_in + qtd11_in

                linhas.append([
                    nome,
                    qtd8_out,
                    qtd11_out,
                    total_entregas,
                    "",  # Coluna PAGO deixa vazia - a validação booleana cria o checkbox
                    pix,
                    None,
                    vale,
                    garantia,
                    qtd8_in,
                    qtd11_in,
                    "",
                    hora_ini,
                    hora_fim
                ])

            end_row = len(linhas)
            # Usa USER_ENTERED para interpretar fórmulas e dados
            print(f"📝 Inserindo {end_row} linhas na planilha '{sheet_title}' (colunas A-N)")
            ws.update(values=linhas, range_name=f"A1:N{end_row}", value_input_option="USER_ENTERED")
            print(f"✅ Dados inseridos com sucesso")

            for idx in range(2, end_row + 1):
                ws.update_cell(idx, 7, f"=B{idx}*8+C{idx}*11-H{idx}+I{idx}")

            requests = []

            # Cores para cada coluna (A-N) no cabeçalho
            cores_colunas = [
                {"red": 0.0, "green": 0.4, "blue": 1.0},      # A - MOTOBOY (Azul vibrante)
                {"red": 0.0, "green": 1.0, "blue": 0.0},      # B -  8 (Verde vibrante)
                {"red": 0.0, "green": 1.0, "blue": 0.0},      # C -  11 (Verde vibrante)
                {"red": 0.0, "green": 1.0, "blue": 1.0},      # D - TOTAL ENTREGAS (Ciano vibrante)
                {"red": 1.0, "green": 0.5, "blue": 0.0},      # E - PAGO (Laranja vibrante)
                {"red": 1.0, "green": 0.0, "blue": 1.0},      # F - PIX (Magenta vibrante)
                {"red": 1.0, "green": 0.0, "blue": 0.0},      # G - TOTAL (Vermelho vibrante)
                {"red": 1.0, "green": 1.0, "blue": 0.0},      # H - Gastos (Amarelo vibrante)
                {"red": 0.0, "green": 0.8, "blue": 0.4},      # I - Garantido (Verde água)
                {"red": 0.0, "green": 0.7, "blue": 0.2},      # J -  8 (Verde floresta)
                {"red": 0.0, "green": 0.7, "blue": 0.2},      # K -  11 (Verde floresta)
                {"red": 0.8, "green": 0.8, "blue": 0.8},      # L - Vazio (Cinza clarão)
                {"red": 1.0, "green": 0.8, "blue": 0.0},      # M - Inicio (Ouro)
                {"red": 1.0, "green": 0.8, "blue": 0.0}       # N - Fim (Ouro)
            ]

            print(f"🎨 Aplicando cores ao cabeçalho ({len(cores_colunas)} colunas)...")
            # Adiciona formatação de cor para cada coluna do cabeçalho
            for col_idx, cor in enumerate(cores_colunas):
                requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws.id,
                            "startRowIndex": 0,
                            "endRowIndex": 1,
                            "startColumnIndex": col_idx,
                            "endColumnIndex": col_idx + 1
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": cor,
                                "textFormat": {
                                    "bold": True,
                                    "fontSize": 11,
                                    "foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0}
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,textFormat)"
                    }
                })

            requests.extend([
                {
                    "updateSheetProperties": {
                        "properties": {
                            "sheetId": ws.id,
                            "gridProperties": {"frozenRowCount": 1}
                        },
                        "fields": "gridProperties.frozenRowCount"
                    }
                }
            ])

            if end_row >= 2:
                requests.append({
                    "setDataValidation": {
                        "range": {
                            "sheetId": ws.id,
                            "startRowIndex": 1,
                            "endRowIndex": end_row,
                            "startColumnIndex": 4,
                            "endColumnIndex": 5
                        },
                        "rule": {
                            "condition": {"type": "BOOLEAN"},
                            "showCustomUi": True
                        }
                    }
                })

                requests.append({
                    "addConditionalFormatRule": {
                        "rule": {
                            "ranges": [
                                {
                                    "sheetId": ws.id,
                                    "startRowIndex": 1,
                                    "endRowIndex": end_row,
                                    "startColumnIndex": 0,
                                    "endColumnIndex": 14
                                }
                            ],
                            "booleanRule": {
                                "condition": {
                                    "type": "CUSTOM_FORMULA",
                                    "values": [{"userEnteredValue": "=$E2=TRUE"}]
                                },
                                "format": {
                                    "backgroundColor": {"red": 0.0, "green": 0.6, "blue": 0.0}
                                }
                            }
                        },
                        "index": 0
                    }
                })

            # Ajusta largura das colunas
            for col_idx in range(14):  # A até N
                if col_idx == 5:  # Coluna PIX (F) - deixa pequena
                    largura = 80
                elif col_idx in (1, 2):  # Entregas 8/11
                    largura = 60
                else:
                    largura = 120  # Colunas normais com largura razoável
                
                requests.append({
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": ws.id,
                            "dimension": "COLUMNS",
                            "startIndex": col_idx,
                            "endIndex": col_idx + 1
                        },
                        "properties": {"pixelSize": largura},
                        "fields": "pixelSize"
                    }
                })

            print(f"📤 Enviando {len(requests)} requisições ao Google Sheets (cores, congelamento, validação, etc)...")
            sh.batch_update({"requests": requests})
            print(f"✅ Tudo atualizado com sucesso!")
            self.mostrar_toast(f"Planilha atualizada: {sheet_title}", "success")
        except Exception as e:
            import traceback
            print(f"❌ ERRO ao atualizar Sheets:")
            print(f"   {type(e).__name__}: {e}")
            traceback.print_exc()
            self.mostrar_toast(f"Erro ao atualizar Sheets: {e}", "error")

    def enviar_fechamento_manual(self):
        try:
            with open(ARQUIVO_COMANDO, 'w', encoding='utf-8') as f:
                f.write("FECHAMENTO_MANUAL")
            self.mostrar_toast("Fechamento enviado ao robo.", "success")
        except Exception as e:
            self.mostrar_toast(f"Erro: {e}", "error")

    def _obter_nome_aba_sheets(self):
        data_str = self.data_var.get().strip()
        if "/" in data_str:
            partes = data_str.split("/")
        else:
            partes = data_str.split("-")
        if len(partes) >= 2:
            return f"{partes[0]}/{partes[1]}"
        return data_str or None

    def obter_motoboys_disponiveis(self):
        motos_do_excel = []
        arq = self._excel_path()

        if os.path.exists(arq):
            try:
                _, df_m = self.carregar_excel_cache(arq)
                if df_m is not None and 'MOTOBOY' in df_m.columns:
                    motos_do_excel = [str(n) for n in df_m['MOTOBOY'].dropna().unique() if str(n).upper() != "RETIRADA"]
            except:
                pass

        motos_cadastrados = list(self.config_data.get("motoboys", {}).values())
        return sorted(set(motos_do_excel + motos_cadastrados))

    def atualizar_lista_motoboys_vales(self):
        if not hasattr(self, "combo_moto_vales"):
            return

        motos = self.obter_motoboys_disponiveis()
        if motos:
            self.combo_moto_vales.configure(values=motos)
            if self.combo_moto_vales.get() not in motos:
                self.combo_moto_vales.set("Selecione um Motoboy")
        else:
            self.combo_moto_vales.configure(values=["Nenhum motoboy encontrado"])
            self.combo_moto_vales.set("Nenhum motoboy encontrado")

    def calcular_total_vales_moto(self, nome):
        arq = self._excel_path()
        if not nome or not os.path.exists(arq):
            return 0.0

        rows = self.carregar_vales_cache(arq)
        if not rows:
            return 0.0

        nome_norm = normalizar_texto(nome)
        total = 0.0
        for _, _, moto, val, _ in rows:
            if normalizar_texto(str(moto)) == nome_norm:
                try:
                    total += float(val or 0.0)
                except:
                    pass
        return total

    # ═════════════════════════════════════════════════════════════════════════════
    # SEÇÃO 4: ABA MONITOR & DASHBOARD
    # ═════════════════════════════════════════════════════════════════════════════
    # Exibição de pedidos do dia com filtro de busca e integrações
    # ─────────────────────────────────────────────────────────────────────────────

    # -------------------------------------------------------------------------
    # ==================================================================================
    #  SEÇÃO 4: ABA MONITOR & DASHBOARD
    # ==================================================================================
    # Responsável por: Exibição do painel em tempo real com pedidos ativos,
    # status de motoboys, e monitoramento de operações do dia.
    # ==================================================================================

    def setup_aba_monitor(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(4, weight=1)

        self.fr_alerta = ctk.CTkFrame(parent, fg_color=COR_ZE_AMARELO, corner_radius=8, height=40)
        ctk.CTkLabel(self.fr_alerta, text="⚠️ NOVO BAIRRO DETECTADO - CADASTRE NA ABA ZONAS", text_color="black", font=FONT_BOLD).place(relx=0.5, rely=0.5, anchor="center")

        self.fr_cards = ctk.CTkFrame(parent, fg_color="transparent")
        self.fr_cards.grid(row=1, column=0, sticky="ew", pady=(0, 20))
        self.fr_cards.grid_columnconfigure((0, 1, 2), weight=1)
        self.fr_cards.grid_columnconfigure(3, weight=0)

        self.card_entregas = self.criar_card_stat(self.fr_cards, "ENTREGAS", "0", COR_ZE_AMARELO, 0)
        self.card_retiradas = self.criar_card_stat(self.fr_cards, "RETIRADAS", "0", "#FF9F0A", 1)
        self.card_fatur = self.criar_card_stat(self.fr_cards, "FATURAMENTO", "R$ 0,00", COR_SUCCESS, 2)

        self.fr_quick_actions = ctk.CTkFrame(self.fr_cards, fg_color=COR_CARD_BG, corner_radius=12, border_width=1, border_color=COR_BORDA)
        self.fr_quick_actions.grid(row=0, column=3, sticky="ns", padx=(10, 0))
        self.fr_quick_actions.grid_columnconfigure(0, weight=1)

        ctk.CTkButton(
            self.fr_quick_actions, text="HISTORICO", command=self.verificar_historico,
            fg_color=COR_CARD_BG, border_width=1, border_color=COR_BORDA,
            text_color=COR_TEXT_SEC, height=30, hover_color="#333", font=("Segoe UI Bold", 12)
        ).grid(row=0, column=0, sticky="ew", padx=8, pady=(10, 6))

        ctk.CTkButton(
            self.fr_quick_actions, text="ESTOQUE", command=self.atualizar_estoque_manual,
            fg_color=COR_CARD_BG, border_width=1, border_color=COR_BORDA,
            text_color=COR_TEXT_SEC, height=30, hover_color="#333", font=("Segoe UI Bold", 12)
        ).grid(row=1, column=0, sticky="ew", padx=8, pady=6)

        ctk.CTkButton(
            self.fr_quick_actions, text="CANCELADAS", command=self.enviar_canceladas,
            fg_color="transparent", border_width=1, border_color=COR_DANGER,
            text_color=COR_DANGER, height=30, hover_color="#2b1111", font=("Segoe UI Bold", 12)
        ).grid(row=2, column=0, sticky="ew", padx=8, pady=(6, 10))

        fr_acoes = ctk.CTkFrame(parent, fg_color=COR_CARD_BG, height=70, corner_radius=12, border_width=1, border_color=COR_BORDA)
        fr_acoes.grid(row=3, column=0, sticky="ew", pady=(0, 20))

        self.ent_busca = ctk.CTkEntry(fr_acoes, placeholder_text="🔍 Buscar...", width=160, height=40, font=FONT_MAIN, border_color=COR_BORDA)
        self.ent_busca.pack(side="left", padx=12, pady=15, fill="x", expand=True)
        self.ent_busca.bind("<KeyRelease>", self.filtrar_tabela_busca)

        ctk.CTkLabel(fr_acoes, text="Motoboy:", font=("Roboto", 11, "bold"), text_color=COR_TEXT_SEC).pack(side="left", padx=(15, 5))
        
        self.combo_motos_ativos = ctk.CTkComboBox(fr_acoes, width=200, height=40, font=FONT_MAIN, border_color=COR_BORDA, state="readonly", values=["Clique para carregar"])
        self.combo_motos_ativos.pack(side="left", padx=5, fill="x")
        self.combo_motos_ativos.set("Clique para carregar")
        # Quando clica, carrega os motoboys
        self.combo_motos_ativos.bind("<Button-1>", self._atualizar_combo_motoboys)

        ctk.CTkButton(
            fr_acoes, text="📋 IMPRIMIR", width=120, height=40, command=self.imprimir_combo_motoboy,
            fg_color=COR_ZE_AMARELO, text_color="black", hover_color=COR_ZE_HOVER, font=("Segoe UI Bold", 12)
        ).pack(side="left", padx=5)

        # Data operacional ao lado do fechamento/imprimir
        self.fr_data_ops = ctk.CTkFrame(fr_acoes, fg_color="transparent")
        self.fr_data_ops.pack(side="left", padx=(10, 0))
        ctk.CTkLabel(
            self.fr_data_ops, text="Data:", font=FONT_MAIN, text_color=COR_TEXT_SEC
        ).pack(side="left", padx=(0, 6))

        if TEM_CALENDARIO:
            self.ent_data = DateEntry(
                self.fr_data_ops, textvariable=self.data_var, width=12,
                background="#222", foreground="white", borderwidth=0,
                date_pattern="dd-mm-yyyy", font=FONT_BOLD
            )
            self.ent_data.pack(side="left")
        else:
            self.ent_data = ctk.CTkEntry(
                self.fr_data_ops, textvariable=self.data_var, justify="center",
                fg_color=COR_CARD_BG, width=120, height=36
            )
            self.ent_data.pack(side="left")

        self.btn_atualizar_data = ctk.CTkButton(
            self.fr_data_ops, text="ATUALIZAR", command=self.carregar_tabela,
            fg_color=COR_CARD_BG, hover_color="#333", border_width=1,
            border_color=COR_BORDA, height=36, font=FONT_BOLD
        )
        self.btn_atualizar_data.pack(side="left", padx=(8, 0))

        # Status de alertas automáticos
        self.lbl_alerta_auto_status = ctk.CTkLabel(
            fr_acoes,
            text="",
            font=FONT_BOLD,
            text_color=COR_TEXT_SEC
        )
        self.lbl_alerta_auto_status.pack(side="right", padx=(10, 20))
        self._atualizar_status_alerta_auto()

        self.tab_tabela = ctk.CTkTabview(
            parent, fg_color="transparent", text_color="black", segmented_button_fg_color=COR_NEON,
            segmented_button_selected_color=COR_ZE_AMARELO, segmented_button_selected_hover_color=COR_ZE_HOVER,
            border_width=1, border_color=COR_NEON
        )
        self.tab_tabela.grid(row=4, column=0, sticky="nsew")
        
        # 3 abas diferentes - Entregas primeiro
        self.t_entrega = self.tab_tabela.add("ENTREGAS")
        self.t_retirada = self.tab_tabela.add("RETIRADAS")
        self.t_cancelado = self.tab_tabela.add("CANCELADOS")

        for tab in [self.t_entrega, self.t_retirada, self.t_cancelado]:
            tab.grid_columnconfigure(0, weight=1)
            tab.grid_rowconfigure(0, weight=1)

        self.tree_entrega = self.criar_tabela_dark(self.t_entrega, ["Hora", "Numero", "Cliente", "Bairro", "Status", "Motoboy", "Valor"])
        self.tree_entrega.bind("<Double-1>", self.ao_clicar_duas_vezes_pedido)
        
        self.tree_retirada = self.criar_tabela_dark(self.t_retirada, ["Hora", "Numero", "Cliente", "Bairro", "Status", "Motoboy", "Valor"])
        self.tree_retirada.bind("<Double-1>", self.ao_clicar_duas_vezes_pedido)
        
        self.tree_cancelado = self.criar_tabela_dark(self.t_cancelado, ["Hora", "Numero", "Cliente", "Bairro", "Status", "Motoboy", "Valor"])
        self.tree_cancelado.bind("<Double-1>", self.ao_clicar_duas_vezes_pedido)
        
        if not self._defer_startup:
            self.carregar_tabela()
        
        # Carrega alertas pendentes ao iniciar
        self.after(500, self.carregar_alertas_atraso)

    def criar_card_stat(self, parent, titulo, valor, cor, col_idx):
        f = ctk.CTkFrame(parent, fg_color=COR_CARD_BG, corner_radius=12, border_width=1, border_color=COR_BORDA)
        f.grid(row=0, column=col_idx, sticky="ew", padx=6)
        ctk.CTkFrame(f, width=5, fg_color=cor, corner_radius=0).pack(side="left", fill="y")
        ctk.CTkLabel(f, text=titulo, text_color=COR_TEXT_SEC, font=("Roboto", 11, "bold")).pack(anchor="w", padx=15, pady=(15, 0))
        lbl = ctk.CTkLabel(f, text=valor, text_color="white", font=("Roboto", 30, "bold"))
        lbl.pack(anchor="w", padx=15, pady=(0, 15))
        return lbl

    def criar_tabela_dark(self, parent, colunas):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background=COR_BG_APP, fieldbackground=COR_BG_APP, foreground="#E0E0E0", rowheight=30, borderwidth=0, font=("Roboto", 9))
        style.configure("Treeview.Heading", background=COR_CARD_BG, foreground="#bbb", relief="flat", font=("Roboto Bold", 9))
        style.map("Treeview", background=[('selected', '#333')], foreground=[('selected', 'white')])

        fr_table = ctk.CTkFrame(parent, fg_color="transparent")

        if isinstance(parent, ctk.CTkTabview) or parent.winfo_name().startswith("!ctktabview"):
            fr_table.pack(fill="both", expand=True)
        else:
            fr_table.grid(row=0, column=0, sticky="nsew")

        tree = ttk.Treeview(fr_table, columns=colunas, show="headings", style="Treeview")
        for c in colunas:
            tree.heading(c, text=c.upper())
            tree.column(c, width=100, anchor="center")

        if "Cliente" in colunas: tree.column("Cliente", width=220, anchor="w")
        if "Bairro" in colunas: tree.column("Bairro", width=160, anchor="w")
        if "PRODUTO" in colunas: tree.column("PRODUTO", width=250, anchor="w")

        scroll_y = ctk.CTkScrollbar(fr_table, command=tree.yview, fg_color="transparent", button_color=COR_BORDA)
        tree.configure(yscroll=scroll_y.set)

        tree.pack(side="left", fill="both", expand=True)
        scroll_y.pack(side="right", fill="y")
        return tree

    # -------------------------------------------------------------------------
    # -------------------------------------------------------------------------
    # ==================================================================================
    #  SEÇÃO 6: ABA VALES & DESCONTOS
    # ==================================================================================
    # Responsável por: Gerenciamento de vales concedidos, descontos aplicados
    # e histórico de ajustes financeiros por motoboy.
    # ==================================================================================

    def setup_aba_vales(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(2, weight=1)

        # ÁREA DE ALERTAS DE ATRASO (PENDÊNCIA)
        self.fr_alertas_atraso = ctk.CTkScrollableFrame(parent, fg_color=COR_CARD_BG, corner_radius=8, border_width=1, border_color=COR_DANGER, height=0)
        self.fr_alertas_atraso.grid(row=0, column=0, sticky="ew", pady=(10, 10), padx=20)
        self.fr_alertas_atraso.grid_remove()  # Esconde inicialmente

        # Cabeçalho
        self.fr_vales_top = ctk.CTkFrame(parent, fg_color="transparent")
        self.fr_vales_top.grid(row=1, column=0, sticky="ew", padx=20, pady=20)
        self.fr_vales_top.grid_columnconfigure(0, weight=1)
        self.fr_vales_top.grid_columnconfigure(1, weight=0)

        self.fr_vales_left = ctk.CTkFrame(self.fr_vales_top, fg_color="transparent")
        self.fr_vales_left.grid(row=0, column=0, sticky="w")

        ctk.CTkLabel(self.fr_vales_left, text="GESTÃO DE VALES E ADIANTAMENTOS", font=FONT_TITLE, text_color=COR_ZE_AMARELO).pack(side="left")

        ctk.CTkLabel(self.fr_vales_left, text="Motoboy:", text_color=COR_TEXT_SEC, font=FONT_MAIN).pack(side="left", padx=(20, 8))
        self.combo_moto_vales = ctk.CTkComboBox(
            self.fr_vales_left, width=240, height=36, font=FONT_MAIN, border_color=COR_BORDA,
            state="readonly", values=["Selecione um Motoboy"]
        )
        self.combo_moto_vales.set("Selecione um Motoboy")
        self.combo_moto_vales.pack(side="left")

        self.fr_vales_actions = ctk.CTkFrame(self.fr_vales_top, fg_color="transparent")
        self.fr_vales_actions.grid(row=0, column=1, sticky="e")

        # Botões de Ação
        ctk.CTkButton(self.fr_vales_actions, text="💰 LANÇAR VALE", command=self.adicionar_vale_manual, fg_color=COR_SUCCESS, text_color="#003300", font=FONT_BOLD).pack(side="left", padx=10)
        ctk.CTkButton(self.fr_vales_actions, text="✏️ EDITAR", command=self.editar_vale, fg_color="#333", text_color="white").pack(side="left", padx=10)
        ctk.CTkButton(self.fr_vales_actions, text="🗑️ EXCLUIR", command=self.excluir_vale, fg_color="transparent", border_width=1, border_color=COR_DANGER, text_color=COR_DANGER).pack(side="left", padx=10)

        # Tabela
        fr_table_container = ctk.CTkFrame(parent, fg_color="transparent")
        fr_table_container.grid(row=2, column=0, sticky="nsew", padx=20)

        colunas = ["ID", "HORA", "MOTOBOY", "VALOR", "MOTIVO"]
        self.tree_vales = self.criar_tabela_dark(fr_table_container, colunas)
        
        # Ajuste de largura das colunas
        self.tree_vales.column("ID", width=0, stretch=False) # ID Oculto (índice da linha no Excel)
        self.tree_vales.column("HORA", width=100, anchor="center")
        self.tree_vales.column("MOTOBOY", width=250, anchor="w")
        self.tree_vales.column("VALOR", width=150, anchor="e")
        self.tree_vales.column("MOTIVO", width=300, anchor="w")

        # Botão Atualizar
        ctk.CTkButton(parent, text="↻ ATUALIZAR LISTA", command=self.carregar_tabela_vales, fg_color=COR_CARD_BG).grid(row=3, column=0, pady=10)

    def carregar_tabela_vales(self):
        if self._loading_vales:
            return
        
        arq = self._excel_path()
        
        # Verificação inteligente: só carrega se mudou
        try:
            mtime = os.path.getmtime(arq)
        except OSError:
            return
        
        # Se arquivo não mudou, não recarrega
        if mtime == getattr(self, 'cache_vales_load_mtime', None):
            return
        
        self._loading_vales = True
        self._set_loading(True)
        self.cache_vales_load_mtime = mtime

        def worker():
            if not os.path.exists(arq):
                data = {"missing": True}
            else:
                rows = self.carregar_vales_cache(arq)
                data = {"rows": rows}
            self._enqueue_ui(lambda d=data: self._render_vales(d))

        threading.Thread(target=worker, daemon=True).start()

    def _render_vales(self, data):
        self._loading_vales = False
        self._set_loading(False)
        
        if data.get("missing"):
            return

        rows = data.get("rows")
        if rows is None:
            self.mostrar_toast("Erro ao ler vales.", "error")
            return

        # OTIMIZAÇÃO: Limpa e insere com desabilitação de redraw
        self.tree_vales.configure(selectmode='none')
        
        children = self.tree_vales.get_children()
        if children:
            self.tree_vales.delete(*children)
        
        # Limita linhas se necessário
        if len(rows) > MAX_ROWS_DISPLAY:
            rows = rows[-MAX_ROWS_DISPLAY:]

        # Insere todos os itens
        if rows:
            for r, hora, moto, val, motivo in rows:
                val_fmt = f"R$ {val:.2f}"
                self.tree_vales.insert("", "end", values=[r, hora, moto, val_fmt, motivo])
        
        # Reabilita seleção
        self.tree_vales.configure(selectmode='browse')

    def adicionar_vale_manual(self):
        # Janela simples para adicionar direto na planilha
        moto = self.combo_moto_vales.get() if hasattr(self, "combo_moto_vales") else ""
        if not moto or "Selecione" in moto or "Nenhum" in moto:
            self.mostrar_toast("Selecione um motoboy na lista.", "error")
            return
        
        valor_str = simpledialog.askstring("Novo Vale", f"Valor para {moto}:")
        if not valor_str: return
        
        motivo = simpledialog.askstring("Novo Vale", "Motivo:") or "Manual Painel"
        
        try:
            val_float = float(valor_str.replace(",", "."))
            
            # Salva direto no Excel usando openpyxl para não depender do robô
            arq = self._excel_path()
            if not os.path.exists(arq): return
            
            wb = openpyxl.load_workbook(arq)
            if "VALES" not in wb.sheetnames:
                wb.create_sheet("VALES")
            ws = wb["VALES"]
            if ws.max_row == 1 and ws["A1"].value is None:
                ws.append(["HORA", "MOTOBOY", "VALOR", "MOTIVO"])
            
            hora = datetime.now().strftime('%H:%M')
            ws.append([hora, moto, val_float, motivo])
            wb.save(arq)
            self.invalidar_cache_excel()
            self.carregar_tabela_vales()
            self.mostrar_toast("Vale adicionado com sucesso!", "success")
            
        except Exception as e:
            self.mostrar_toast(f"Erro ao salvar: {e}", "error")

    def excluir_vale(self):
        sel = self.tree_vales.selection()
        if not sel:
            self.mostrar_toast("Selecione um vale para excluir.", "error")
            return
            
        dados = self.tree_vales.item(sel[0], "values")
        linha_excel = int(dados[0]) # Recupera o ID oculto (número da linha)
        
        if messagebox.askyesno("Confirmar", f"Excluir vale de {dados[2]} ({dados[3]})?"):
            try:
                arq = self._excel_path()
                wb = openpyxl.load_workbook(arq)
                ws = wb["VALES"]
                
                ws.delete_rows(linha_excel)
                wb.save(arq)
                self.invalidar_cache_excel()
                self.carregar_tabela_vales()
                self.mostrar_toast("Vale excluído!", "success")
            except Exception as e:
                self.mostrar_toast(f"Erro ao excluir: {e}", "error")

    def editar_vale(self):
        sel = self.tree_vales.selection()
        if not sel: return
        
        dados = self.tree_vales.item(sel[0], "values")
        linha_excel = int(dados[0])
        nome_atual = dados[2]
        valor_atual = dados[3].replace("R$ ", "").replace(".", "") # Limpa formatação visual
        
        novo_valor = simpledialog.askstring("Editar", f"Novo valor para {nome_atual}:", initialvalue=valor_atual)
        if novo_valor:
            try:
                val_float = float(novo_valor.replace(",", "."))
                arq = self._excel_path()
                wb = openpyxl.load_workbook(arq)
                ws = wb["VALES"]
                
                # Atualiza a coluna 3 (Valor) da linha específica
                ws.cell(row=linha_excel, column=3).value = val_float
                wb.save(arq)
                self.invalidar_cache_excel()
                self.carregar_tabela_vales()
                self.mostrar_toast("Vale atualizado!", "success")
            except Exception as e:
                self.mostrar_toast(f"Erro ao editar: {e}", "error")
    
    # ==================================================================================
    #  SEÇÃO 7: ABA ESTOQUE
    # ==================================================================================
    # Responsável por: Controle de inventário, monitoramento de itens,
    # alertas de estoque baixo e atualização de quantidades.
    # ==================================================================================

    class AbaEstoque(ctk.CTkFrame):
        def __init__(self, parent):
            super().__init__(parent)
            self.estoque_data = []
        
            self.setup_aba_estoque(self)
            self.estoque_data = self.carregar_estoque()
            self.atualizar_tabela_estoque()

    def setup_aba_estoque(self, parent):
        # Configuração do Grid Principal da Aba
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(3, weight=1)

        # --- 1. Título ---
        ctk.CTkLabel(parent, text="CONTROLE DE ESTOQUE (AUTOMÁTICO)", font=FONT_TITLE, text_color=COR_ZE_AMARELO).grid(row=0, column=0, pady=(20, 10))

        # --- 2. Área de Busca ---
        fr_busca = ctk.CTkFrame(parent, fg_color="transparent")
        fr_busca.grid(row=1, column=0, sticky="ew", padx=30, pady=(0, 10))

        self.ent_busca = ctk.CTkEntry(fr_busca, placeholder_text="🔍 Pesquisar produto...", height=38, fg_color=COR_CARD_BG, border_color=COR_BORDA)
        self.ent_busca.pack(side="left", fill="x", expand=True, padx=(0, 12))
        self.ent_busca.bind("<Return>", lambda _: self.comando_buscar())

        ctk.CTkButton(fr_busca, text="BUSCAR", command=self.comando_buscar, width=110, height=38, fg_color=COR_CARD_BG, hover_color=COR_BORDA).pack(side="right")

        # --- 3. Frame de Inputs (Adicionar/Editar) ---
        fr_inputs = ctk.CTkFrame(parent, fg_color=COR_CARD_BG, height=65)
        fr_inputs.grid(row=2, column=0, sticky="ew", padx=30, pady=12)

        self.ent_prod = ctk.CTkEntry(fr_inputs, placeholder_text="Nome do Produto", width=260, height=42)
        self.ent_prod.pack(side="left", padx=12, pady=12)

        self.ent_preco = ctk.CTkEntry(fr_inputs, placeholder_text="R$ Preço", width=110, height=42)
        self.ent_preco.pack(side="left", padx=7, pady=12)

        self.ent_qtd = ctk.CTkEntry(fr_inputs, placeholder_text="Qtd", width=90, height=42)
        self.ent_qtd.pack(side="left", padx=7, pady=12)

        self.combo_cat = ctk.CTkComboBox(fr_inputs, values=self.ESTOQUE_CATEGORIAS, width=210, height=42, state="readonly")
        self.combo_cat.set("Selecione a Categoria")
        self.combo_cat.pack(side="left", padx=7, pady=12)

        ctk.CTkButton(fr_inputs, text="💾 SALVAR", command=self.add_produto, fg_color=COR_SUCCESS, text_color="#003300", font=FONT_BOLD, width=110).pack(side="left", padx=12)
        ctk.CTkButton(fr_inputs, text="🧹 LIMPAR", command=self.limpar_campos, fg_color="transparent", border_width=1, border_color=COR_BORDA, text_color="silver", width=90).pack(side="left", padx=7)

        ctk.CTkButton(fr_inputs, text="🗑️", command=self.del_produto, fg_color="transparent", border_width=1, border_color=COR_DANGER, text_color=COR_DANGER, width=50).pack(side="right", padx=12)
        ctk.CTkButton(fr_inputs, text="🛒 LISTA DE COMPRAS", command=self.gerar_lista_compras, fg_color="#E91E63", hover_color="#C2185B", font=FONT_BOLD, width=160).pack(side="right", padx=12)

        # --- 4. Tabela (Treeview) ---
        fr_tabela_container = ctk.CTkFrame(parent, fg_color="transparent")
        fr_tabela_container.grid(row=3, column=0, sticky="nsew", padx=30, pady=12)

        colunas = ["PRODUTO", "NIVEL", "STATUS", "PRECO", "FORNECEDOR"]

        style = ttk.Style()
        style.configure("Treeview", rowheight=32, font=("Arial", 11))
        style.configure("Treeview.Heading", font=("Arial", 11, "bold"))

        self.tree_estoque = ttk.Treeview(fr_tabela_container, columns=colunas, show="headings", style="Treeview", selectmode="browse")

        self.tree_estoque.heading("PRODUTO", text="PRODUTO", anchor="w")
        self.tree_estoque.heading("NIVEL", text="NÍVEL VISUAL", anchor="w")
        self.tree_estoque.heading("STATUS", text="QTD / STATUS", anchor="center")
        self.tree_estoque.heading("PRECO", text="PREÇO", anchor="e")
        self.tree_estoque.heading("FORNECEDOR", text="FORNECEDOR", anchor="w")

        self.tree_estoque.column("PRODUTO", width=320, minwidth=220)
        self.tree_estoque.column("NIVEL", width=110, minwidth=110)
        self.tree_estoque.column("STATUS", width=130, anchor="center")
        self.tree_estoque.column("PRECO", width=110, anchor="e")
        self.tree_estoque.column("FORNECEDOR", width=170, anchor="w")

        sc_y = ctk.CTkScrollbar(fr_tabela_container, command=self.tree_estoque.yview, fg_color="transparent", button_color=COR_BORDA)
        self.tree_estoque.configure(yscroll=sc_y.set)

        self.tree_estoque.pack(side="left", fill="both", expand=True)
        sc_y.pack(side="right", fill="y")

        self.tree_estoque.bind("<<TreeviewSelect>>", self.ao_selecionar_item)
    # ================= LÓGICA DE DADOS =================

    def carregar_estoque(self):
        """Carrega o estoque do arquivo JSON, garantindo retorno de lista."""
        try:
            if os.path.exists("estoque.json"):
                with open("estoque.json", "r", encoding="utf-8") as f:
                    dados = json.load(f)
                    # Conversão de legado (Dict -> Lista)
                    if isinstance(dados, dict):
                        lista_convertida = []
                        for nome, qtd in dados.items():
                            lista_convertida.append({
                                "nome": nome,
                                "estoque_fisico": qtd,
                                "categoria": "GERAL",
                                "preco_venda": 0.0,
                                "fornecedor": "-"
                            })
                        return lista_convertida
                    return dados
            return []
        except (FileNotFoundError, json.JSONDecodeError):
            return []

    def salvar_estoque_disk(self):
        """Salva a lista de estoque no disco."""
        try:
            with open("estoque.json", 'w', encoding='utf-8') as f:
                json.dump(self.estoque_data, f, indent=4, ensure_ascii=False)
        except:
            pass

    def comando_buscar(self):
        """Acionado pelo botão buscar ou Enter"""
        termo = self.ent_busca.get()
        self.atualizar_tabela_estoque(filtro=termo)

    def limpar_campos(self):
        """Limpa os inputs e tira a seleção da tabela"""
        self.ent_prod.delete(0, 'end')
        self.ent_preco.delete(0, 'end')
        self.ent_qtd.delete(0, 'end')
        self.combo_cat.set("Selecione a Categoria")
        if self.tree_estoque.selection():
            self.tree_estoque.selection_remove(self.tree_estoque.selection())

    def ao_selecionar_item(self, _):
        """Preenche os campos quando clica na tabela"""
        selecao = self.tree_estoque.selection()
        if not selecao: return

        item = self.tree_estoque.item(selecao[0])
        tags = item['tags']
        
        # Se clicou na categoria (pastinha), ignora
        if 'categoria' in tags:
            return

        valores = item['values']
        # values[0] é o nome com espaços visuais. Limpamos para buscar.
        nome_sujo = valores[0].strip() 
        
        # Busca o item real na lista de dados para ter precisão nos números
        item_real = next((i for i in self.estoque_data if i.get("nome", "").lower() == nome_sujo.lower()), None)
        
        if item_real:
            self.ent_prod.delete(0, 'end')
            self.ent_prod.insert(0, item_real.get("nome", "").title())
            
            self.ent_preco.delete(0, 'end')
            self.ent_preco.insert(0, str(item_real.get("preco_venda", 0.0)))
            
            self.ent_qtd.delete(0, 'end')
            self.ent_qtd.insert(0, str(item_real.get("estoque_fisico", 0)))

            categoria = item_real.get("categoria") or self.identificar_categoria(item_real.get("nome", ""))
            self.combo_cat.set(categoria)

    def add_produto(self):
        """Adiciona ou Atualiza produto."""
        nome = self.ent_prod.get().strip().lower()
        qtd_str = self.ent_qtd.get().strip()
        preco_str = self.ent_preco.get().strip().replace(',', '.')
        categoria_selecionada = self.combo_cat.get()
        
        if "Selecione" in categoria_selecionada:
            categoria = self.identificar_categoria(nome)
        else:
            categoria = categoria_selecionada
        
        if nome and qtd_str:
            try:
                qtd = float(qtd_str) if '.' in qtd_str else int(qtd_str)
                preco = float(preco_str) if preco_str else 0.0
                
                # Verifica se já existe para atualizar
                encontrado = False
                for item in self.estoque_data:
                    if item.get("nome", "").lower() == nome:
                        item["estoque_fisico"] = qtd # Atualiza Qtd
                        item["preco_venda"] = preco  # Atualiza Preço
                        item["nome"] = nome # Garante que o nome seja salvo em minúsculo
                        item["categoria"] = categoria # Atualiza a categoria
                        encontrado = True
                        break
                
                if not encontrado:
                    self.estoque_data.append({
                        "nome": nome,
                        "estoque_fisico": qtd,
                        "categoria": categoria,
                        "preco_venda": preco,
                        "fornecedor": "Manual"
                    })

                self.salvar_estoque_disk()
                self.atualizar_tabela_estoque()
                self.limpar_campos()
                self.mostrar_toast("Produto salvo!", "success")
            except ValueError:
                self.mostrar_toast("Erro: Preço ou Qtd inválidos", "error")

    def del_produto(self):
        """Remove produto selecionado da lista."""
        nome = self.ent_prod.get().strip().lower()
        if not nome:
            self.mostrar_toast("Selecione um produto!", "error")
            return
            
        self.estoque_data = [item for item in self.estoque_data if item.get("nome", "").lower() != nome]
        
        self.salvar_estoque_disk()
        self.atualizar_tabela_estoque()
        self.limpar_campos()
        self.mostrar_toast("Produto removido!", "success")

    # ================= VISUAL E UTILITÁRIOS =================

    def gerar_barra_visual(self, atual, maximo=100):
        tamanho_barra = 10
        if maximo == 0: maximo = 1
        divisor = maximo if atual <= maximo else atual
        if divisor == 0: divisor = 1
        percentual = min(atual / divisor, 1.0)
        blocos_cheios = int(tamanho_barra * percentual)
        blocos_vazios = tamanho_barra - blocos_cheios
        return f"{'█' * blocos_cheios}{'░' * blocos_vazios}"

    def identificar_categoria(self, nome_produto):
        nome = str(nome_produto).lower()
        regras = {
            "🍺 CERVEJAS": ["skol", "brahma", "antarctica", "budweiser", "heineken", "spaten", "corona", "original", "bohemia", "polar", "subzero", "serrana", "bavaria", "kaiser", "proibida", "becks", "stella", "eisenbahn", "michelob"],
            "🍸 DESTILADOS": ["vodka", "whisky", "gin", "cachaça", "rum", "tequila", "licor", "campari", "aperol", "velho barreiro", "51", "smirnoff", "absolut", "jack", "red label", "white horse", "passport", "ballantines", "chivas", "conhaque", "dreher", "sakerita", "old parr", "montilla", "malibu"],
            "🥤 NÃO ALCOÓLICOS": ["coca", "pepsi", "fanta", "guaraná", "sprite", "sukita", "soda", "agua", "água", "h2oh", "gatorade", "suco", "del valle", "tonica", "refrigerante"],
            "⚡ ENERGÉTICOS": ["red bull", "monster", "baly", "vibe", "tnt"],
            "🍷 VINHOS & ICES": ["vinho", "cantina", "pergola", "pérgola", "sangue de boi", "chalise", "canção", "catuaba", "ice", "beats", "syn", "galiotto", "dom bosco", "cider"],
            "🍟 MERCEARIA & DIVERSOS": ["gelo", "carvao", "carvão", "salgadinho", "batata", "amendoim", "doritos", "ruffles", "cheetos", "fandangos", "chocolate", "halls", "trident", "seda", "cigarro", "fumo", "isqueiro", "copo", "baconzitos", "cebolitos", "torcida", "kit kat"]
        }
        for categoria, palavras in regras.items():
            if any(p in nome for p in palavras): return categoria
        return "📦 OUTROS"

    def atualizar_tabela_estoque(self, filtro=""):
        # Limpa tabela visualmente
        for item in self.tree_estoque.get_children(): 
            self.tree_estoque.delete(item)
        
        # Import local para evitar erro se não estiver no topo
        import unicodedata 

        def normalizar_estoque_nome(nome):
            if not nome: return ""
            try:
                nfkd = unicodedata.normalize('NFKD', str(nome))
                t = "".join([c for c in nfkd if not unicodedata.combining(c)]).lower().strip()
                for ch in "-_/()[]{}.,;:": t = t.replace(ch, " ")
                return " ".join(t.split())
            except: return str(nome).lower().strip()

        # Termos ignorados na busca
        termos_ignorar = ["vasilhame incluso", "pack 12", "pack12", "pack 18", "pack18"]
        termos_ignorar_norm = [normalizar_estoque_nome(t) for t in termos_ignorar]

        # Organização por Categoria
        estoque_organizado = {}
        termo_busca = filtro.lower() if filtro else ""

        palavras_filtro = normalizar_estoque_nome(termo_busca).split()
        for item in self.estoque_data:
            nome_raw = item.get("nome", "Desconhecido")
            nome_norm = normalizar_estoque_nome(nome_raw)
            # Filtro de Busca: exige todas as palavras do filtro no nome
            if palavras_filtro and not all(p in nome_norm for p in palavras_filtro):
                continue
            # Pula packs se necessário
            if any(t in nome_norm for t in termos_ignorar_norm):
                continue
            # Organiza
            cat = item.get("categoria", self.identificar_categoria(nome_raw))
            if cat not in estoque_organizado: estoque_organizado[cat] = []
            estoque_organizado[cat].append(item)

        # Renderiza na Tabela
        for categoria in sorted(estoque_organizado.keys()):
            nome_cat = categoria.upper()
            id_pai = self.tree_estoque.insert("", "end", values=[f"📂 {nome_cat}", "", "", "", ""], open=True, tags=("categoria",))
            
            # Produtos ordenados por nome
            for item in sorted(estoque_organizado[categoria], key=lambda x: x.get("nome", "")):
                nome = item.get("nome", "").title()
                try:
                    qtd = float(item.get("estoque_fisico", 0))
                    preco = float(item.get("preco_venda", 0))
                except: qtd = 0; preco = 0

                forn = item.get("fornecedor", "-")

                # Regras de Status
                if qtd <= 0: status = "🚨 CRÍTICO"; tag = "critico"; limite = 20
                elif qtd <= 3: status = "⚠️ BAIXO"; tag = "baixo"; limite = 50
                else: status = "🔋 ESTÁVEL"; tag = "normal"; limite = 100

                barra = self.gerar_barra_visual(qtd, limite)
                qtd_visual = int(qtd) if qtd.is_integer() else f"{qtd:.2f}"
                preco_visual = f"R$ {preco:.2f}"

                self.tree_estoque.insert(id_pai, "end", 
                                       values=[f"   {nome}", barra, f"{qtd_visual} | {status}", preco_visual, forn], 
                                       tags=(tag,))

        # Cores das Tags
        self.tree_estoque.tag_configure('categoria', background="#111", foreground=COR_ZE_AMARELO, font=("Roboto", 10, "bold"))
        self.tree_estoque.tag_configure('critico', foreground=COR_DANGER, font=("Consolas", 10, "bold"))
        self.tree_estoque.tag_configure('baixo', foreground="#FFD600", font=("Consolas", 10))
        self.tree_estoque.tag_configure('normal', foreground="#00E676", font=("Consolas", 10))

    def gerar_lista_compras(self):
        """Gera arquivo de texto com itens críticos."""
        if not self.estoque_data:
            self.mostrar_toast("Estoque vazio!", "error")
            return

        ignore_list_str = self.config_data.get("compras_ignore_list", "")
        ignore_list = [item.strip().lower() for item in ignore_list_str.split(',') if item.strip()]

        compras = {}
        itens_criticos = 0
        
        # Lógica simplificada de agrupamento
        for item in self.estoque_data:
            try:
                qtd = float(item.get("estoque_fisico", 0))
                if qtd <= 3:
                    nome_original = item.get("nome", "")
                    if nome_original.strip().lower() in ignore_list:
                        continue

                    forn = item.get("fornecedor", "OUTROS").upper()
                    nome = nome_original.title()
                    if forn not in compras: compras[forn] = []
                    compras[forn].append((nome, qtd))
                    itens_criticos += 1
            except: continue

        if itens_criticos == 0:
            self.mostrar_toast("Estoque cheio! Nada para comprar.", "success")
            return

        nome_arquivo = f"Lista_Compras_{datetime.now().strftime('%d-%m-%Y')}.txt"
        try:
            with open(nome_arquivo, "w", encoding="utf-8") as f:
                f.write(f"🛒 LISTA DE REPOSIÇÃO - ZÉ DELIVERY\n📅 {datetime.now().strftime('%d/%m/%Y')}\n\n")
                for forn, itens in compras.items():
                    f.write(f"🏢 {forn}\n" + "-"*30 + "\n")
                    for nome, qtd in itens:
                        f.write(f"[ ] {nome:<30} (Est.: {qtd})\n")
                    f.write("\n")
            
            self.mostrar_toast(f"Lista gerada: {nome_arquivo}", "success")
            os.startfile(nome_arquivo)
        except Exception as e:
            self.mostrar_toast(f"Erro: {e}", "error")

    def mostrar_toast(self, mensagem, tipo="info"):
        # Se você tiver uma função global de toast, use-a.
        # Caso contrário, usa messagebox simples
        try:
            from tkinter import messagebox
            if tipo == "error": messagebox.showerror("Aviso", mensagem)
            else: messagebox.showinfo("Info", mensagem)
        except: pass
    # ==================================================================================
    #  SEÇÃO 9: ABA CONFIG
    # ==================================================================================
    # Responsável por: Configurações do sistema como horários, limites,
    # integração com APIs e preferências do usuário.
    # ==================================================================================

    def setup_aba_config(self, parent):
        fr = ctk.CTkFrame(parent, fg_color=COR_CARD_BG, corner_radius=15, border_width=1, border_color=COR_BORDA)
        fr.place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(fr, text="CONFIGURAÇÕES GERAIS", font=FONT_TITLE).pack(pady=20, padx=60)

        self.ent_email = ctk.CTkEntry(fr, placeholder_text="E-mail Zé", width=350)
        self.ent_email.pack(pady=5)
        self.ent_email.insert(0, self.config_data.get("email_ze", ""))

        self.ent_senha = ctk.CTkEntry(fr, placeholder_text="Senha Zé", show="●", width=350)
        self.ent_senha.pack(pady=5)
        self.ent_senha.insert(0, self.config_data.get("senha_ze", ""))

        ctk.CTkLabel(fr, text="📱 Telegram", font=FONT_BOLD, text_color=COR_ZE_AMARELO).pack(pady=(15, 5))

        self.ent_tele_token = ctk.CTkEntry(fr, placeholder_text="Token Bot", width=350)
        self.ent_tele_token.pack(pady=5)
        self.ent_tele_token.insert(0, self.config_data.get("telegram_token", ""))

        self.ent_tele_chat = ctk.CTkEntry(fr, placeholder_text="Seu Chat ID", width=350)
        self.ent_tele_chat.pack(pady=5)
        self.ent_tele_chat.insert(0, self.config_data.get("telegram_chat_id", ""))

        ctk.CTkLabel(fr, text="🧾 Google Sheets", font=FONT_BOLD, text_color=COR_ZE_AMARELO).pack(pady=(15, 5))

        sheets_cfg = self.config_data.get("google_sheets", {})
        self.ent_sheets_id = ctk.CTkEntry(fr, placeholder_text="Sheets ID", width=350)
        self.ent_sheets_id.pack(pady=5)
        self.ent_sheets_id.insert(0, sheets_cfg.get("sheets_id", ""))

        self.ent_sheets_json = ctk.CTkEntry(fr, placeholder_text="Service Account JSON (caminho)", width=350)
        self.ent_sheets_json.pack(pady=5)
        self.ent_sheets_json.insert(0, sheets_cfg.get("service_account_json", ""))

        ctk.CTkLabel(fr, text="🛒 Lista de Compras", font=FONT_BOLD, text_color=COR_ZE_AMARELO).pack(pady=(15, 5))
        
        self.ent_compras_ignore = ctk.CTkEntry(fr, placeholder_text="Itens a ignorar, separados por vírgula", width=350)
        self.ent_compras_ignore.pack(pady=5)
        self.ent_compras_ignore.insert(0, self.config_data.get("compras_ignore_list", ""))

        ctk.CTkLabel(fr, text="📢 WhatsApp Alertas", font=FONT_BOLD, text_color=COR_ZE_AMARELO).pack(pady=(15, 5))

        fr_alerta_auto = ctk.CTkFrame(fr, fg_color="transparent")
        fr_alerta_auto.pack(pady=5)

        ctk.CTkLabel(fr_alerta_auto, text="Alertas automáticos:", font=FONT_MAIN, text_color=COR_TEXT_SEC).pack(side="left", padx=(0, 10))

        self.switch_alerta_auto = ctk.CTkSwitch(
            fr_alerta_auto,
            text="",
            width=50,
            onvalue=True,
            offvalue=False
        )
        self.switch_alerta_auto.pack(side="left")

        alerta_auto_ativo = self.config_data.get("alerta_retirada_auto", False)
        if alerta_auto_ativo:
            self.switch_alerta_auto.select()
        else:
            self.switch_alerta_auto.deselect()

        fr_mencao = ctk.CTkFrame(fr, fg_color="transparent")
        fr_mencao.pack(pady=5)
        
        ctk.CTkLabel(fr_mencao, text="Mencionar nas mensagens:", font=FONT_MAIN, text_color=COR_TEXT_SEC).pack(side="left", padx=(0, 10))
        
        self.switch_mencao = ctk.CTkSwitch(
            fr_mencao, 
            text="", 
            width=50,
            onvalue=True,
            offvalue=False
        )
        self.switch_mencao.pack(side="left")
        
        # Carrega estado atual
        mencao_ativa = self.config_data.get("whatsapp_mencao_ativa", False)
        if mencao_ativa:
            self.switch_mencao.select()
        else:
            self.switch_mencao.deselect()

        ctk.CTkButton(fr, text="SALVAR TUDO", command=self.salvar_creds, width=350, fg_color=COR_SUCCESS, text_color="#003300", font=FONT_BOLD).pack(pady=20)

        ctk.CTkLabel(fr, text="☁️ BACKUP", font=FONT_BOLD, text_color=COR_ZE_AMARELO).pack(pady=(10, 5))
        self.lbl_path_backup = ctk.CTkLabel(fr, text=self.config_data.get("path_backup", "Nenhuma pasta"), text_color=COR_TEXT_SEC, font=("Consolas", 10))
        self.lbl_path_backup.pack()

        fr_bkp = ctk.CTkFrame(fr, fg_color="transparent")
        fr_bkp.pack(pady=10)
        ctk.CTkButton(fr_bkp, text="PASTA", command=self.selecionar_pasta_backup, width=100).pack(side="left", padx=5)
        ctk.CTkButton(fr_bkp, text="FAZER BACKUP", command=self.fazer_backup, width=150, fg_color="#333").pack(side="left", padx=5)

    def salvar_creds(self):
        try:
            self.config_data["email_ze"] = self.ent_email.get()
            self.config_data["senha_ze"] = self.ent_senha.get()
            self.config_data["telegram_token"] = self.ent_tele_token.get()
            self.config_data["telegram_chat_id"] = self.ent_tele_chat.get()
            self.config_data.setdefault("google_sheets", {})["sheets_id"] = self.ent_sheets_id.get().strip()
            self.config_data.setdefault("google_sheets", {})["service_account_json"] = self.ent_sheets_json.get().strip()
            self.config_data["compras_ignore_list"] = self.ent_compras_ignore.get()
            self.config_data["alerta_retirada_auto"] = self.switch_alerta_auto.get()
            self.config_data["whatsapp_mencao_ativa"] = self.switch_mencao.get()
            self.salvar_config()
            self.google_sheets_config = self._carregar_google_sheets_config()
            self._atualizar_status_alerta_auto()
            self.mostrar_toast("Salvo com sucesso!", "success")
        except Exception as e:
            import traceback
            print("Erro em salvar_creds:")
            traceback.print_exc()
            try:
                self.mostrar_toast(f"Erro ao salvar: {e}", "error")
            except Exception:
                pass

    def selecionar_pasta_backup(self):
        p = filedialog.askdirectory()
        if p:
            self.config_data["path_backup"] = p
            self.lbl_path_backup.configure(text=p)
            self.salvar_config()

    def fazer_backup(self):
        dest = self.config_data.get("path_backup", "")
        if not dest or not os.path.exists(dest):
            self.mostrar_toast("Pasta inválida", "error")
            return
        try:
            arq_excel = self._excel_path()
            if os.path.exists(arq_excel): shutil.copy2(arq_excel, os.path.join(dest, os.path.basename(arq_excel)))
            if os.path.exists(ARQUIVO_CONFIG): shutil.copy2(ARQUIVO_CONFIG, os.path.join(dest, "backup_config.json"))
            if os.path.exists(ARQUIVO_ESTOQUE): shutil.copy2(ARQUIVO_ESTOQUE, os.path.join(dest, "backup_estoque.json"))
            self.mostrar_toast("Backup OK!", "success")
        except Exception as e:
            self.mostrar_toast(f"Erro: {e}", "error")

    # ==================================================================================
    #  SEÇÃO 10: ABA LOGS
    # ==================================================================================
    # Responsável por: Exibição de logs de operação, eventos do sistema,
    # erros e histórico de atividades para debugging.
    # ==================================================================================

    def setup_aba_logs(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(parent, text=">_ TERMINAL DO SISTEMA", font=FONT_BOLD, text_color=COR_ZE_AMARELO).grid(row=0, column=0, sticky="w", pady=(0, 10))

        self.txt_logs = ctk.CTkTextbox(parent, font=FONT_MONO, text_color="#00FF9C", fg_color="#000", corner_radius=6)
        self.txt_logs.grid(row=1, column=0, sticky="nsew", pady=10)
        self.txt_logs.configure(state="disabled")

        fr_input = ctk.CTkFrame(parent, fg_color=COR_CARD_BG, height=50)
        fr_input.grid(row=2, column=0, sticky="ew", pady=10)

        self.ent_cmd = ctk.CTkEntry(fr_input, placeholder_text="Comando...", height=40, font=FONT_MAIN, border_width=0, fg_color="transparent")
        self.ent_cmd.pack(side="left", fill="x", expand=True, padx=10)
        self.ent_cmd.bind("<Return>", lambda _: self.enviar_comando_robo())

        ctk.CTkButton(fr_input, text="ENVIAR", command=self.enviar_comando_robo, height=35, fg_color="#333").pack(side="right", padx=10)

    # ==================================================================================
    #  SEÇÃO 11: ABA MOTOS
    # ==================================================================================
    # Responsável por: Gerenciamento de motoboys, status em tempo real,
    # performance individual e histórico de entregas.
    # ==================================================================================

    def setup_aba_motos(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(parent, text="EQUIPE DE ENTREGAS", font=FONT_TITLE).grid(row=0, column=0, pady=20)

        fr_list = ctk.CTkFrame(parent, fg_color=COR_CARD_BG, border_width=1, border_color=COR_BORDA)
        fr_list.grid(row=1, column=0, sticky="nsew", padx=40, pady=10)

        self.lst_motos = tk.Listbox(fr_list, bg=COR_CARD_BG, fg="white", bd=0, highlightthickness=0, font=("Roboto", 12), selectbackground="#333")
        self.lst_motos.pack(fill="both", expand=True, padx=20, pady=20)

        fr_btns = ctk.CTkFrame(parent, fg_color="transparent")
        fr_btns.grid(row=2, column=0, sticky="ew", padx=40, pady=20)

        ctk.CTkButton(fr_btns, text="+ ADICIONAR", command=self.add_moto, fg_color="#333", height=45).pack(side="left", expand=True, padx=5, fill="x")
        ctk.CTkButton(fr_btns, text="REMOVER", command=self.del_moto, fg_color="transparent", border_width=1, border_color=COR_DANGER, text_color=COR_DANGER, height=45).pack(side="left", expand=True, padx=5, fill="x")
        ctk.CTkButton(fr_btns, text="💾 SALVAR", command=self.salvar_motos_disk, fg_color=COR_SUCCESS, text_color="#003300", height=45, font=FONT_BOLD).pack(side="left", expand=True, padx=5, fill="x")

        self.atualizar_lista_motos()

    # ==================================================================================
    #  SEÇÃO 12: ABA PIX
    # ==================================================================================
    # Responsável por: Cadastro de chaves PIX para motoboys.
    # ==================================================================================

    def setup_aba_pix(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(3, weight=1)

        ctk.CTkLabel(parent, text="CADASTRO DE PIX", font=FONT_TITLE).grid(row=0, column=0, pady=20)

        fr_top = ctk.CTkFrame(parent, fg_color=COR_CARD_BG, border_width=1, border_color=COR_BORDA)
        fr_top.grid(row=1, column=0, sticky="ew", padx=40, pady=10)
        fr_top.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(fr_top, text="Motoboy:", text_color=COR_TEXT_SEC, font=FONT_BOLD).grid(row=0, column=0, padx=15, pady=15, sticky="w")
        self.combo_pix_moto = ctk.CTkComboBox(
            fr_top, width=240, height=36, font=FONT_MAIN, border_color=COR_BORDA,
            values=[], command=lambda _: self._on_pix_moto_select()
        )
        self.combo_pix_moto.grid(row=0, column=1, padx=10, pady=15, sticky="ew")
        self.combo_pix_moto.set("Selecione um Motoboy")

        ctk.CTkLabel(fr_top, text="PIX:", text_color=COR_TEXT_SEC, font=FONT_BOLD).grid(row=0, column=2, padx=15, pady=15, sticky="w")
        self.ent_pix_valor = ctk.CTkEntry(fr_top, height=36, font=FONT_MAIN, border_color=COR_BORDA, fg_color=COR_CARD_BG)
        self.ent_pix_valor.grid(row=0, column=3, padx=10, pady=15, sticky="ew")
        fr_top.grid_columnconfigure(3, weight=2)

        fr_actions = ctk.CTkFrame(parent, fg_color="transparent")
        fr_actions.grid(row=2, column=0, sticky="ew", padx=40, pady=(0, 10))
        ctk.CTkButton(fr_actions, text="💾 SALVAR", command=self.salvar_pix_motoboy, fg_color=COR_SUCCESS, text_color="#003300", height=45, font=FONT_BOLD).pack(side="left", expand=True, padx=5, fill="x")
        ctk.CTkButton(fr_actions, text="LIMPAR", command=self._limpar_pix_form, fg_color="#333", height=45).pack(side="left", expand=True, padx=5, fill="x")

        fr_list = ctk.CTkFrame(parent, fg_color=COR_CARD_BG, border_width=1, border_color=COR_BORDA)
        fr_list.grid(row=3, column=0, sticky="nsew", padx=40, pady=(0, 20))
        fr_list.grid_rowconfigure(0, weight=1)
        fr_list.grid_columnconfigure(0, weight=1)

        self.lst_pix = tk.Listbox(fr_list, bg=COR_CARD_BG, fg="white", bd=0, highlightthickness=0, font=("Roboto", 12), selectbackground="#333")
        self.lst_pix.grid(row=0, column=0, sticky="nsew", padx=20, pady=15)
        self.lst_pix.bind("<<ListboxSelect>>", self._on_pix_list_select)

        self._pix_list_keys = []
        self.atualizar_lista_pix()

    def atualizar_lista_pix(self):
        if not hasattr(self, "combo_pix_moto"):
            return
        motos = self.obter_motoboys_disponiveis()
        if motos:
            self.combo_pix_moto.configure(values=motos)
            if self.combo_pix_moto.get() not in motos:
                self.combo_pix_moto.set("Selecione um Motoboy")
        else:
            self.combo_pix_moto.configure(values=["Nenhum motoboy encontrado"])
            self.combo_pix_moto.set("Nenhum motoboy encontrado")

        if not hasattr(self, "lst_pix"):
            return
        self.lst_pix.delete(0, "end")
        pix_map = self.config_data.get("pix_motoboys", {})
        self._pix_list_keys = []
        for nome in sorted(pix_map.keys(), key=lambda s: str(s).lower()):
            pix = pix_map.get(nome, "")
            self._pix_list_keys.append(nome)
            self.lst_pix.insert("end", f"{nome} | {pix}")

    def _on_pix_moto_select(self):
        nome = self.combo_pix_moto.get() if hasattr(self, "combo_pix_moto") else ""
        if not nome or "Selecione" in nome or "Nenhum" in nome:
            return
        pix = self._obter_pix_motoboy(nome)
        if hasattr(self, "ent_pix_valor"):
            self.ent_pix_valor.delete(0, "end")
            if pix:
                self.ent_pix_valor.insert(0, pix)

    def _on_pix_list_select(self, _):
        if not hasattr(self, "lst_pix"):
            return
        sel = self.lst_pix.curselection()
        if not sel:
            return
        idx = sel[0]
        if idx >= len(self._pix_list_keys):
            return
        nome = self._pix_list_keys[idx]
        if hasattr(self, "combo_pix_moto"):
            self.combo_pix_moto.set(nome)
        pix = self._obter_pix_motoboy(nome)
        if hasattr(self, "ent_pix_valor"):
            self.ent_pix_valor.delete(0, "end")
            if pix:
                self.ent_pix_valor.insert(0, pix)

    def _limpar_pix_form(self):
        if hasattr(self, "combo_pix_moto"):
            self.combo_pix_moto.set("Selecione um Motoboy")
        if hasattr(self, "ent_pix_valor"):
            self.ent_pix_valor.delete(0, "end")

    def salvar_pix_motoboy(self):
        nome = self.combo_pix_moto.get() if hasattr(self, "combo_pix_moto") else ""
        if not nome or "Selecione" in nome or "Nenhum" in nome:
            self.mostrar_toast("Selecione um motoboy.", "error")
            return
        pix = self.ent_pix_valor.get().strip() if hasattr(self, "ent_pix_valor") else ""
        if not pix:
            self.mostrar_toast("Digite a chave PIX.", "error")
            return
        self.config_data.setdefault("pix_motoboys", {})[nome] = pix
        self.salvar_config()
        self.atualizar_lista_pix()
        self.mostrar_toast("PIX salvo!", "success")

    # ==================================================================================
    #  SEÇÃO 12: ABA BAIRROS
    # ==================================================================================
    # Responsável por: Gerenciamento de bairros/zonas, valores de entrega,
    # controle de cobertura e délimitação de áreas.
    # ==================================================================================

    def setup_aba_bairros(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(2, weight=1)

        ctk.CTkLabel(parent, text="GESTÃO DE ZONAS & TAXAS", font=FONT_TITLE).grid(row=0, column=0, pady=20)

        fr_top = ctk.CTkFrame(parent, fg_color="transparent")
        fr_top.grid(row=1, column=0, sticky="ew", padx=40)

        self.ent_bairro = ctk.CTkEntry(fr_top, placeholder_text="Nome do Bairro...", height=45, font=FONT_MAIN, border_color=COR_BORDA, fg_color=COR_CARD_BG)
        self.ent_bairro.pack(side="left", fill="x", expand=True, padx=(0, 15))

        ctk.CTkButton(fr_top, text="+ TAXA R$ 8,00", command=lambda: self.add_bairro(8), height=45, width=140, fg_color="#333").pack(side="left", padx=5)
        ctk.CTkButton(fr_top, text="+ TAXA R$ 11,00", command=lambda: self.add_bairro(11), height=45, width=140, fg_color="#444").pack(side="left", padx=5)

        fr_cols = ctk.CTkFrame(parent, fg_color="transparent")
        fr_cols.grid(row=2, column=0, sticky="nsew", padx=40, pady=25)
        fr_cols.grid_columnconfigure((0, 1), weight=1)
        fr_cols.grid_rowconfigure(0, weight=1)

        # Coluna 8
        f8 = ctk.CTkFrame(fr_cols, fg_color=COR_CARD_BG, border_width=1, border_color=COR_BORDA)
        f8.grid(row=0, column=0, sticky="nsew", padx=10)
        ctk.CTkLabel(f8, text="ZONA 1 (R$ 8,00)", text_color="white", font=FONT_BOLD).pack(pady=15)
        self.lst_8 = tk.Listbox(f8, bg=COR_CARD_BG, fg="white", bd=0, highlightthickness=0, font=("Roboto", 11), selectbackground="#333")
        self.lst_8.pack(fill="both", expand=True, padx=15)
        ctk.CTkButton(f8, text="Remover", command=lambda: self.del_bairro(8), fg_color="transparent", text_color=COR_TEXT_SEC, hover_color="#333").pack(fill="x", padx=15, pady=10)

        # Coluna 11
        f11 = ctk.CTkFrame(fr_cols, fg_color=COR_CARD_BG, border_width=1, border_color=COR_BORDA)
        f11.grid(row=0, column=1, sticky="nsew", padx=10)
        ctk.CTkLabel(f11, text="ZONA 2 (R$ 11,00)", text_color="white", font=FONT_BOLD).pack(pady=15)
        self.lst_11 = tk.Listbox(f11, bg=COR_CARD_BG, fg="white", bd=0, highlightthickness=0, font=("Roboto", 11), selectbackground="#333")
        self.lst_11.pack(fill="both", expand=True, padx=15)
        ctk.CTkButton(f11, text="Remover", command=lambda: self.del_bairro(11), fg_color="transparent", text_color=COR_TEXT_SEC, hover_color="#333").pack(fill="x", padx=15, pady=10)

        ctk.CTkButton(parent, text="💾 APLICAR ALTERAÇÕES", command=self.salvar_bairros_disk, height=50, fg_color=COR_SUCCESS, text_color="#003300", font=FONT_BOLD).grid(row=3, column=0, sticky="ew", padx=50, pady=(0, 30))
        self.atualizar_listas_bairros()

    # ================= FUNÇÕES DE SISTEMA =================
    def mostrar_toast(self, mensagem, tipo="info"):
        if not hasattr(self, "frame_toast"):
            return
        cor = COR_SUCCESS if tipo == "success" else COR_DANGER if tipo == "error" else COR_ZE_AMARELO
        cor_texto = "black" if tipo != "error" else "white"
        self.frame_toast.configure(fg_color=cor)
        self.lbl_toast.configure(text=mensagem, text_color=cor_texto)
        self.frame_toast.place(relx=0.5, rely=0.05, anchor="n")
        self.after(3000, lambda: self.frame_toast.place_forget())

    def _enqueue_ui(self, fn):
        if not hasattr(self, "_ui_queue"):
            return
        self._ui_queue.put(fn)

    # ==================================================================================
    #  SEÇÃO 13: SISTEMA DE CACHE (OTIMIZAÇÕES)
    # ==================================================================================
    # Responsável por: Implementação das 5 otimizações de performance:
    # mtime checking, selective columns, auto-refresh, TreeView opt, Pandas cache.
    # ==================================================================================

    def _process_ui_queue(self):
        if not self.winfo_exists():
            return
        processed = 0
        max_batch = 5  # Processa até 5 itens por vez para evitar travamento
        while processed < max_batch:
            try:
                fn = self._ui_queue.get_nowait()
            except queue.Empty:
                break
            try:
                fn()
                processed += 1
            except Exception:
                pass
        delay = UI_QUEUE_INTERVAL_MS if processed > 0 else UI_QUEUE_IDLE_MS
        self.after(delay, self._process_ui_queue)

    def buscar_robo_no_sistema(self):
        now = time.time()
        if now - getattr(self, "_last_robo_check_ts", 0) < 5:
            return
        self._last_robo_check_ts = now
        try:
            cmd = 'wmic process where "name=\'python.exe\'" get commandline'
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, shell=True, startupinfo=startupinfo)
            out, _ = proc.communicate()
            if out and "robo.py" in out:
                self.robo_rodando = True
                self.btn_power.configure(text="PARAR OPERAÇÃO", fg_color=COR_CARD_BG, border_color=COR_DANGER, text_color=COR_DANGER, hover_color="#2b1111")
                self.lbl_status_text.configure(text="SISTEMA ONLINE", text_color=COR_SUCCESS)
                self.lbl_status_dot.configure(text_color=COR_SUCCESS)
                self.log_sistema("⚠️ Sessão detectada! Sincronizado.")
                if not self.processo_robo:
                    self.iniciar_tail_log()
        except Exception as e:
            self.log_sistema(f"Erro ao escanear processos: {e}")

    def controlar_janela(self, acao):
        alvos = []
        try:
            def enum_handler(hwnd, results):
                if win32gui.IsWindowVisible(hwnd):
                    title = win32gui.GetWindowText(hwnd)
                    if ("Google Chrome" in title or "Chrome" in title) and "ZÉ DELIVERY" not in title:
                        results.append((hwnd, title))
                elif acao == "show":
                    title = win32gui.GetWindowText(hwnd)
                    if ("Google Chrome" in title or "Chrome" in title) and "ZÉ DELIVERY" not in title:
                        results.append((hwnd, title))

            win32gui.EnumWindows(enum_handler, alvos)
            if not alvos:
                self.mostrar_toast("Nenhum Chrome encontrado", "error")
                return
            for hwnd, _ in alvos:
                if acao == "hide":
                    win32gui.ShowWindow(hwnd, win32con.SW_HIDE)
                elif acao == "show":
                    win32gui.ShowWindow(hwnd, win32con.SW_SHOW)
                    win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
        except:
            pass

    def toggle_robo(self):
        if not self.robo_rodando:
            self.iniciar_robo()
        else:
            self.parar_robo()

    def iniciar_robo(self):
        if not os.path.exists("robo.py"):
            messagebox.showerror("Erro", "robo.py não encontrado!")
            return
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        env_dict = os.environ.copy()
        env_dict["PYTHONIOENCODING"] = "utf-8"
        try:
            self.processo_robo = subprocess.Popen(
                ["python", "-u", "robo.py"], stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                stdin=subprocess.PIPE, text=True, encoding='utf-8', errors='ignore',
                creationflags=subprocess.CREATE_NO_WINDOW, bufsize=1, env=env_dict
            )
            self.log_tail_running = False
            self.robo_rodando = True
            self.btn_power.configure(text="PARAR OPERAÇÃO", fg_color=COR_CARD_BG, border_color=COR_DANGER, text_color=COR_DANGER, hover_color="#2b1111")
            self.lbl_status_text.configure(text="SISTEMA ONLINE", text_color=COR_SUCCESS)
            self.lbl_status_dot.configure(text_color=COR_SUCCESS)
            threading.Thread(target=self.ler_output_robo, daemon=True).start()
            self.mudar_aba("logs")
            self.mostrar_toast("Sistema Iniciado", "success")
        except Exception as e:
            messagebox.showerror("Erro", str(e))
            self.parar_robo()

    def parar_robo(self):
        if self.processo_robo:
            self.processo_robo.terminate()
            self.processo_robo = None
        else:
            subprocess.call("wmic process where \"commandline like '%robo.py%'\" call terminate", shell=True, stderr=subprocess.DEVNULL, stdout=subprocess.DEVNULL)
        self.robo_rodando = False
        self.log_tail_running = False
        self.btn_power.configure(text="INICIAR SISTEMA", fg_color=COR_CARD_BG, border_color=COR_SUCCESS, text_color=COR_SUCCESS, hover_color="#1a332a")
        self.lbl_status_text.configure(text="SISTEMA OFFLINE", text_color="#555")
        self.lbl_status_dot.configure(text_color="#333")
        self.log_sistema("🛑 Operação finalizada.")

    def ler_output_robo(self):
        while self.robo_rodando and self.processo_robo:
            try:
                line = self.processo_robo.stdout.readline()
                if not line: break
                self.fila_logs.put(line)
            except:
                break
        if self.robo_rodando:
            self._enqueue_ui(self.parar_robo)

    def iniciar_tail_log(self):
        if self.log_tail_running:
            return
        self.log_tail_running = True
        threading.Thread(target=self.ler_log_arquivo, daemon=True).start()

    def ler_log_arquivo(self):
        try:
            with open(self.log_file_path, "r", encoding="utf-8", errors="ignore") as f:
                try:
                    f.seek(0, os.SEEK_END)
                    size = f.tell()
                    f.seek(max(size - 4096, 0))
                    linhas = f.read().splitlines()
                    for linha in linhas[-100:]:
                        self.fila_logs.put(linha + "\n")
                except Exception:
                    pass

                f.seek(0, os.SEEK_END)
                while self.robo_rodando and not self.processo_robo:
                    linha = f.readline()
                    if linha:
                        self.fila_logs.put(linha)
                    else:
                        time.sleep(0.2)
        finally:
            self.log_tail_running = False

    def atualizar_logs_interface(self):
        # Drena a fila em memória
        count = 0
        max_drain = 100  # Limita quantas linhas processa por vez
        while not self.fila_logs.empty() and count < max_drain:
            self._log_buffer.append(self.fila_logs.get())
            count += 1

        # Só renderiza se a aba de logs está ativa
        if self._aba_atual != "logs":
            # Limita o buffer para não crescer infinitamente
            if len(self._log_buffer) > 500:
                self._log_buffer = self._log_buffer[-500:]
            self.after(LOGS_REFRESH_IDLE_MS, self.atualizar_logs_interface)
            return

        # Renderiza apenas se tem conteúdo novo
        if self._log_buffer:
            self.txt_logs.configure(state="normal")
            
            # Limita o tamanho total do widget de texto (performance)
            current_lines = int(self.txt_logs.index('end-1c').split('.')[0])
            if current_lines > 1000:
                self.txt_logs.delete("1.0", "500.0")  # Remove metade mais antiga
            
            self.txt_logs.insert("end", "".join(self._log_buffer))
            self.txt_logs.see("end")
            self.txt_logs.configure(state="disabled")
            self._log_buffer.clear()

        self.after(LOGS_REFRESH_ACTIVE_MS, self.atualizar_logs_interface)

    def enviar_comando_robo(self):
        cmd = self.ent_cmd.get().strip()
        if cmd and self.robo_rodando:
            try:
                self.processo_robo.stdin.write(cmd + "\n")
                self.processo_robo.stdin.flush()
                self.log_sistema(f">>> ENVIADO: {cmd}")
                self.ent_cmd.delete(0, "end")
            except:
                pass

    def log_sistema(self, msg):
        self.fila_logs.put(f"\n{msg}\n")

    def carregar_config(self):
        if not os.path.exists(ARQUIVO_CONFIG):
            return {"email_ze": "", "senha_ze": "", "motoboys": {}, "bairros": {}, "respostas_bot": {}}
        try:
            with open(ARQUIVO_CONFIG, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            # MIGRAÇÃO AUTOMÁTICA: Converte formato antigo para novo
            if "bairros_8" in config or "bairros_11" in config:
                if "bairros" not in config:
                    config["bairros"] = {}
                
                # Migra bairros_8 para o novo formato
                for bairro in config.get("bairros_8", []):
                    config["bairros"][bairro.lower()] = 8.00
                
                # Migra bairros_11 para o novo formato
                for bairro in config.get("bairros_11", []):
                    config["bairros"][bairro.lower()] = 11.00
                
                # Remove as chaves antigas (opcional - mantém retrocompatibilidade comentado)
                # del config["bairros_8"]
                # del config["bairros_11"]
                
                print("✅ Bairros migrados para novo formato!")
            
            return config
        except:
            return {}

    def salvar_config(self):
        try:
            with open(ARQUIVO_CONFIG, 'w', encoding='utf-8') as f:
                json.dump(self.config_data, f, indent=4, ensure_ascii=False)
            with open(ARQUIVO_COMANDO, 'w', encoding='utf-8') as f:
                f.write("RECARREGAR_CONFIG")
            self.atualizar_cache_bairros()
        except Exception as e:
            import traceback
            print("Erro ao salvar config:")
            traceback.print_exc()
            try:
                self.mostrar_toast(f"Erro ao salvar config: {e}", "error")
            except Exception:
                pass

    def atualizar_cache_bairros(self):
        # Usa o novo formato de dicionário de bairros
        bairros_dict = self.config_data.get("bairros", {})
        self.bairros_conhecidos = set(normalizar_texto(b) for b in bairros_dict.keys())

    def _excel_path(self, data_str=None):
        data_str = (data_str or self.data_var.get()).strip()
        nome = f"Controle_Financeiro_{data_str}.xlsx"
        return os.path.join(get_caminho_base(), nome)

    # --- LÓGICA DO DASHBOARD ---
    def invalidar_cache_excel(self):
        self.cache_detalhe_df = None
        self.cache_motos_df = None
        self.cache_excel_path = None
        self.cache_excel_mtime = None
        self.cache_vales_df = None
        self.cache_vales_path = None
        self.cache_vales_mtime = None

    def carregar_excel_cache(self, arq):
        # Lazy-import pandas to avoid startup overhead when GUI opens
        try:
            import pandas as pd
        except Exception:
            pd = None

        try:
            mtime = os.path.getmtime(arq)
        except OSError:
            return None, None

        if arq == self.cache_excel_path and mtime == self.cache_excel_mtime:
            return self.cache_detalhe_df, self.cache_motos_df

        # Carregar apenas colunas necessárias para melhor performance
        cols_detalhe = [
            'Numero', 'Cliente', 'Bairro', 'Valor (R$)', 'Status', 'Motoboy', 'Hora'
        ]
        cols_motos = ['MOTOBOY', 'QTD TOTAL', 'QTD R$ 8,00', 'QTD R$ 11,00', 'TOTAL A PAGAR (R$)']

        try:
            df = pd.read_excel(
                arq, 
                sheet_name="EXTRATO DETALHADO",
                usecols=lambda col: any(c in col for c in cols_detalhe) if col else False,
                dtype={'Numero': str}
            )
        except Exception:
            try:
                # Fallback: se seleção de colunas falhar, carregar tudo
                df = pd.read_excel(arq, sheet_name="EXTRATO DETALHADO")
            except Exception:
                df = None

        try:
            df_m = pd.read_excel(
                arq,
                sheet_name="PAGAMENTO_MOTOBOYS",
                usecols=lambda col: any(c in col for c in cols_motos) if col else False
            )
        except Exception:
            try:
                df_m = pd.read_excel(arq, sheet_name="PAGAMENTO_MOTOBOYS")
            except Exception:
                df_m = None

        self.cache_excel_path = arq
        self.cache_excel_mtime = mtime
        self.cache_detalhe_df = df
        self.cache_motos_df = df_m
        return df, df_m

    def carregar_vales_cache(self, arq):
        try:
            mtime = os.path.getmtime(arq)
        except OSError:
            return None

        if arq == self.cache_vales_path and mtime == self.cache_vales_mtime:
            return self.cache_vales_df

        rows = []

        def _norm_header(v):
            return str(v or "").strip().upper().replace(" ", "")

        def _eh_header_row(h0, h1):
            return _norm_header(h0) == "HORA" and _norm_header(h1) == "MOTOBOY"

        def _parse_vale_valor(raw):
            if raw is None:
                return 0.0
            if isinstance(raw, (int, float)):
                return float(raw)
            s = str(raw).strip()
            if not s:
                return 0.0
            s = s.replace("R$", "").replace("r$", "").strip()
            if "," in s:
                s = s.replace(".", "").replace(",", ".")
            try:
                return float(s)
            except Exception:
                return 0.0

        def _carregar_vales_openpyxl():
            import openpyxl
            wb = openpyxl.load_workbook(arq, data_only=True)
            if "VALES" not in wb.sheetnames:
                return []
            ws = wb["VALES"]

            first_row_idx = None
            for r, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
                if any(cell not in (None, "") for cell in row):
                    first_row_idx = r
                    first_row = row
                    break

            if first_row_idx is None:
                return []

            start_row = first_row_idx + 1 if _eh_header_row(first_row[0], first_row[1]) else first_row_idx

            for r, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                try:
                    if not row or not row[1] or _eh_header_row(row[0], row[1]):
                        continue
                    hora = row[0]
                    moto = row[1]
                    val = _parse_vale_valor(row[2] if len(row) > 2 else None)
                    motivo = row[3] if len(row) > 3 else ""
                    rows.append((r, hora, moto, val, motivo))
                except Exception:
                    continue
            return rows

        try:
            # Tenta usar pandas primeiro (mais rápido)
            try:
                import pandas as pd
            except Exception:
                pd = None

            if pd is not None:
                df_vales = pd.read_excel(arq, sheet_name="VALES", header=None)
            else:
                df_vales = None

            if df_vales is not None and len(df_vales) > 0:
                for r, row in df_vales.iterrows():
                    try:
                        hora = row[0] if len(row) > 0 else None
                        moto = row[1] if len(row) > 1 else None
                        if not moto or _eh_header_row(hora, moto):
                            continue
                        val = _parse_vale_valor(row[2] if len(row) > 2 else None)
                        motivo = row[3] if len(row) > 3 else ""
                        rows.append((r, hora, moto, val, motivo))
                    except:
                        continue
            else:
                rows = _carregar_vales_openpyxl()
        except Exception:
            try:
                rows = _carregar_vales_openpyxl()
            except Exception:
                rows = None

        self.cache_vales_path = arq
        self.cache_vales_mtime = mtime
        self.cache_vales_df = rows
        return rows

    def carregar_tabela(self, filtro=None):
        if self._loading_monitor:
            return
        
        arq = self._excel_path()
        
        # Verificação inteligente: só carrega se mudou
        try:
            mtime = os.path.getmtime(arq)
        except OSError:
            return
        
        # Se arquivo não mudou, não recarrega (mas processa filtro se houver)
        if mtime == self.cache_monitor_mtime and not filtro:
            return
        
        self._loading_monitor = True
        self._set_loading(True)
        self.cache_monitor_mtime = mtime

        def worker():
            try:
                data = self._coletar_dados_tabela(arq, filtro)
            except Exception as e:
                data = {"error": f"Erro ao carregar dados: {e}"}
            self._enqueue_ui(lambda d=data: self._render_tabela(d))

        threading.Thread(target=worker, daemon=True).start()

    def _coletar_dados_tabela(self, arq, filtro):
        if not os.path.exists(arq):
            return {"missing": True}

        try:
            df, df_m = self.carregar_excel_cache(arq)
        except Exception as e:
            return {"error": f"Erro ao ler planilha: {e}"}

        if df is None:
            return {"error": "Erro ao ler planilha."}

        rows = []
        rows_retirada = []
        rows_cancelado = []
        rows_entrega = []
        val_tot = 0.0
        count_ret = 0
        count_ent = 0
        tem_bairro_desc = False
        ativos_motos = set()

        for _, row in df.iterrows():
            try:
                val = float(row.get('Valor (R$)', 0.0) or 0.0)
            except Exception:
                val = 0.0
            status = str(row.get('Status', '')).upper()
            bairro = normalizar_texto(str(row.get('Bairro', '')))
            moto = str(row.get('Motoboy', ''))

            moto_norm = moto.strip()
            if moto_norm and moto_norm.lower() not in ["nan", "none", "null"] and moto_norm.upper() not in ["RETIRADA", "DESCONHECIDO", "AGUARDANDO..."]:
                ativos_motos.add(moto_norm)

            if filtro:
                cliente = str(row.get('Cliente', ''))
                numero = str(row.get('Numero', ''))
                text_row = f"{cliente} {numero} {bairro} {moto}".lower()
                if filtro not in text_row:
                    continue

            is_cancelado = "CANCEL" in status or "ABANDONED" in status
            is_retirada = val == 0 and not is_cancelado
            is_entrega = not is_cancelado and val > 0

            if not is_cancelado:
                val_tot += val
                if is_retirada:
                    count_ret += 1
                else:
                    count_ent += 1

            tags = []
            if is_cancelado:
                tags.append('cancelado')
            elif is_retirada:
                tags.append('retirada')
            elif bairro not in self.bairros_conhecidos and moto != "RETIRADA":
                tags.append('bairro_desc')
                tem_bairro_desc = True

            item_dict = {
                "values": [
                    row.get('Hora', ''),
                    row.get('Numero', ''),
                    row.get('Cliente', ''),
                    row.get('Bairro', ''),
                    row.get('Status', ''),
                    row.get('Motoboy', ''),
                    f"R$ {val:.2f}"
                ],
                "tags": tuple(tags)
            }
            
            rows.append(item_dict)
            if is_retirada:
                rows_retirada.append(item_dict)
            elif is_cancelado:
                rows_cancelado.append(item_dict)
            elif is_entrega:
                rows_entrega.append(item_dict)

        pagamentos = None
        motos_todos = []
        motos_rows = []
        pagamentos_erro = None

        if df_m is not None:
            try:
                col_qtd8 = "QTD R$ 8,00"
                col_qtd11 = "QTD R$ 11,00"
                col_total = "TOTAL A PAGAR (R$)"

                if col_qtd8 not in df_m.columns or col_qtd11 not in df_m.columns or col_total not in df_m.columns:
                    pagamentos_erro = "Planilha de pagamentos com colunas inesperadas."
                    # Mesmo com erro, preenche motos_todos com os ativos
                    motos_todos = sorted(ativos_motos)
                else:
                    pagamentos = {}
                    for _, r in df_m.iterrows():
                        nome = str(r.get('MOTOBOY', '')).strip()
                        if not nome or nome.upper() == "RETIRADA":
                            continue
                        pagamentos[nome] = {
                            "qtd": int(r.get('QTD TOTAL', 0) or 0),
                            "qtd8": int(r.get(col_qtd8, 0) or 0),
                            "qtd11": int(r.get(col_qtd11, 0) or 0),
                            "total": float(r.get(col_total, 0) or 0.0)
                        }

                    motos_todos = sorted(set(list(pagamentos.keys()) + list(ativos_motos)))
                    for nome in motos_todos:
                        dados = pagamentos.get(nome, {"qtd": 0, "qtd8": 0, "qtd11": 0, "total": 0.0})
                        motos_rows.append([
                            nome,
                            dados['qtd'],
                            dados['qtd8'],
                            dados['qtd11'],
                            f"R$ {dados['total']:.2f}"
                        ])
            except Exception as e:
                pagamentos_erro = f"Erro ao ler pagamentos: {e}"
                # Mesmo com exception, preenche motos_todos com ativos
                if not motos_todos:
                    motos_todos = sorted(ativos_motos)
        else:
            # df_m é None, então preenche motos_todos com ativos encontrados nos pedidos
            motos_todos = sorted(ativos_motos)

        # Fallback: se não achou na planilha de pagamentos, usa os ativos encontrados
        if not motos_todos and ativos_motos:
            motos_todos = sorted(ativos_motos)
        
        # Último fallback: se mesmo assim tiver vazio, usa todos de config.json
        if not motos_todos and self.config_data:
            motos_todos = sorted(self.config_data.get("motoboys", {}).values())
        
        # Se motos_rows ficou vazio mas temos motos_todos, preenche com dados placeholder
        if not motos_rows and motos_todos:
            for nome in motos_todos:
                motos_rows.append([
                    nome,
                    0,  # QTD TOTAL
                    0,  # QTD R$ 8
                    0,  # QTD R$ 11
                    "R$ 0.00"  # A PAGAR
                ])

        return {
            "rows": rows,
            "rows_retirada": rows_retirada,
            "rows_cancelado": rows_cancelado,
            "rows_entrega": rows_entrega,
            "val_tot": val_tot,
            "count_ret": count_ret,
            "count_ent": count_ent,
            "tem_bairro_desc": tem_bairro_desc,
            "missing": False,
            "pagamentos_erro": pagamentos_erro,
            "motos_todos": motos_todos,
            "motos_rows": motos_rows
        }

    def _render_tabela(self, data):
        self._loading_monitor = False
        self._set_loading(False)
        
        if data.get("missing"):
            self.card_entregas.configure(text="0")
            self.card_retiradas.configure(text="0")
            self.card_fatur.configure(text="R$ 0,00")
            return

        if data.get("error"):
            self.mostrar_toast(data["error"], "error")
            return

        # Limpa 3 tables e insere dados separados por tipo
        for tree in [self.tree_retirada, self.tree_cancelado, self.tree_entrega]:
            children = tree.get_children()
            if children:
                tree.delete(*children)
        
        # Insere dados em cada tabela
        rows_retirada = data.get("rows_retirada", [])
        if rows_retirada:
            for item in rows_retirada:
                self.tree_retirada.insert("", "end", values=item["values"], tags=item["tags"])
        
        rows_cancelado = data.get("rows_cancelado", [])
        if rows_cancelado:
            for item in rows_cancelado:
                self.tree_cancelado.insert("", "end", values=item["values"], tags=item["tags"])
        
        rows_entrega = data.get("rows_entrega", [])
        if rows_entrega:
            for item in rows_entrega:
                self.tree_entrega.insert("", "end", values=item["values"], tags=item["tags"])
        
        # Configura tags para todos
        for tree in [self.tree_retirada, self.tree_cancelado, self.tree_entrega]:
            tree.configure(selectmode='browse')
            tree.tag_configure('cancelado', foreground=COR_DANGER)
            tree.tag_configure('retirada', foreground=COR_ZE_AMARELO)
            tree.tag_configure('bairro_desc', background=COR_ZE_AMARELO, foreground="black")

        self.card_entregas.configure(text=str(data.get("count_ent", 0)))
        self.card_retiradas.configure(text=str(data.get("count_ret", 0)))
        self.card_fatur.configure(text=f"R$ {data.get('val_tot', 0.0):.2f}")

        if data.get("tem_bairro_desc"):
            self.fr_alerta.pack(fill="x", pady=5, before=self.fr_cards)
        else:
            self.fr_alerta.pack_forget()

        if data.get("pagamentos_erro"):
            self.mostrar_toast(data["pagamentos_erro"], "error")
        
        # Popula combo com motoboys que já foram coletados em _coletar_dados_tabela
        motos_ativos = data.get("motos_todos", [])
        if motos_ativos:
            self.combo_motos_ativos.configure(values=["Selecione..."] + motos_ativos)
            self.combo_motos_ativos.set("Selecione...")

    def atualizar_lista_motos(self):
        self.lst_motos.delete(0, tk.END)
        for e, n in self.config_data.get("motoboys", {}).items():
            self.lst_motos.insert(tk.END, f"{n}  |  {e}")

    def add_moto(self):
        e = simpledialog.askstring("Novo", "Email:")
        if e:
            n = simpledialog.askstring("Novo", "Nome:")
        if n:
            self.config_data["motoboys"][e.strip().lower()] = n
            self.atualizar_lista_motos()

    def del_moto(self):
        if self.lst_motos.curselection():
            del self.config_data["motoboys"][self.lst_motos.get(self.lst_motos.curselection()[0]).split("|")[1].strip()]
            self.atualizar_lista_motos()

    def salvar_motos_disk(self):
        self.salvar_config()
        self.mostrar_toast("Equipe Salva!", "success")

    def atualizar_listas_bairros(self):
        self.lst_8.delete(0, tk.END)
        self.lst_11.delete(0, tk.END)
        
        # Carrega do dicionário unificado de bairros
        bairros_dict = self.config_data.get("bairros", {})
        
        # Separa por valor e exibe nas listas
        for bairro, valor in sorted(bairros_dict.items()):
            if valor == 8.00 or valor == 8:
                self.lst_8.insert(tk.END, bairro.upper())
            elif valor == 11.00 or valor == 11:
                self.lst_11.insert(tk.END, bairro.upper())

    def add_bairro(self, v):
        # Pega o texto normalizado (minúsculo para consistência)
        n = self.ent_bairro.get().strip().lower()
        
        if n:
            # Garante que o dicionário de bairros existe
            if "bairros" not in self.config_data:
                self.config_data["bairros"] = {}
            
            # Adiciona o bairro com o valor (float)
            self.config_data["bairros"][n] = float(v)
            self.atualizar_listas_bairros()
            self.ent_bairro.delete(0, "end")
            self.mostrar_toast(f"Bairro '{n.upper()}' → R${v} adicionado!", "success")

    def del_bairro(self, v):
        # Seleciona a lista certa (8 ou 11)
        l = self.lst_8 if v == 8 else self.lst_11
        
        # Verifica se tem algo selecionado na lista visual
        if l.curselection():
            item = l.get(l.curselection()[0]).lower()  # Normaliza para minúsculo
            
            # Remove do dicionário de bairros
            if "bairros" in self.config_data and item in self.config_data["bairros"]:
                del self.config_data["bairros"][item]
                self.atualizar_listas_bairros()
                self.mostrar_toast(f"Bairro '{item.upper()}' removido!", "info")

    def salvar_bairros_disk(self):
        self.salvar_config()
        self.mostrar_toast("Zonas Salvas!", "success")

    def filtrar_tabela_busca(self, _):
        termo = self.ent_busca.get().lower()
        if self.search_after_id:
            try:
                self.after_cancel(self.search_after_id)
            except Exception:
                pass

        if not termo:
            self.search_after_id = self.after(150, lambda: self.carregar_tabela())
            return
        self.search_after_id = self.after(250, lambda: self.carregar_tabela(filtro=termo))

    def ao_clicar_duas_vezes_pedido(self, event):
        tree = event.widget
        item_id = tree.selection()
        if not item_id: return
        valores = tree.item(item_id, "values")
        dados_atuais = {
            'Hora': valores[0], 'Numero': valores[1], 'Cliente': valores[2],
            'Bairro': valores[3], 'Status': valores[4], 'Motoboy': valores[5], 'Valor': valores[6]
        }
        JanelaEdicao(self, dados_atuais, self.salvar_alteracao_excel)

    def salvar_alteracao_excel(self, numero_pedido, novos_dados):
        arq = self._excel_path()
        if not os.path.exists(arq): return
        try:
            import pandas as pd
            import openpyxl
            df = pd.read_excel(arq, sheet_name="EXTRATO DETALHADO")
            df['Numero'] = df['Numero'].astype(str)
            index = df.index[df['Numero'] == str(numero_pedido)].tolist()
            if index:
                i = index[0]
                df.at[i, 'Bairro'] = novos_dados['Bairro']
                df.at[i, 'Motoboy'] = novos_dados['Motoboy']
                df.at[i, 'Status'] = novos_dados['Status']
                try:
                    val_float = float(novos_dados['Valor (R$)'])
                except:
                    val_float = 0.0
                df.at[i, 'Valor (R$)'] = val_float
                with pd.ExcelWriter(arq, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name="EXTRATO DETALHADO", index=False)
                self.invalidar_cache_excel()
                self.mostrar_toast(f"Pedido {numero_pedido} Atualizado!", "success")
                self.carregar_tabela()
            else:
                self.mostrar_toast("Pedido não encontrado.", "error")
        except Exception as e:
            self.mostrar_toast(f"Erro ao salvar: {e}", "error")

    def _atualizar_combo_motoboys(self, _=None):
        """Carrega motoboys quando clica no combo"""
        try:
            arq = self._excel_path()
            if not os.path.exists(arq):
                self.combo_motos_ativos.configure(values=["Arquivo não encontrado"])
                return
            
            df, _ = self.carregar_excel_cache(arq)
            if df is None:
                self.combo_motos_ativos.configure(values=["Erro ao ler"])
                return
            
            ativos_motos = set()
            for _, row in df.iterrows():
                moto = str(row.get('Motoboy', '')).strip()
                if moto and moto.lower() not in ["nan", "none", "null"] and moto.upper() not in ["RETIRADA", "DESCONHECIDO", "AGUARDANDO..."]:
                    ativos_motos.add(moto)
            
            if ativos_motos:
                self.combo_motos_ativos.configure(values=["Selecione..."] + sorted(ativos_motos))
                self.combo_motos_ativos.set("Selecione...")
            else:
                self.combo_motos_ativos.configure(values=["Nenhum ativo hoje"])
                self.combo_motos_ativos.set("Nenhum ativo hoje")
        except Exception:
            self.combo_motos_ativos.configure(values=["Erro"])

    def imprimir_combo_motoboy(self):
        nome_selecionado = self.combo_motos_ativos.get()
        if nome_selecionado and "Selecione" not in nome_selecionado:
            with open(ARQUIVO_COMANDO, 'w', encoding='utf-8') as f:
                f.write(f"IMPRIMIR:{nome_selecionado}|{self.data_var.get()}")
            self.mostrar_toast(f"Imprimindo: {nome_selecionado}", "success")
        else:
            self.mostrar_toast("Selecione um motoboy!", "error")

    def enviar_canceladas(self):
        with open(ARQUIVO_COMANDO, 'w', encoding='utf-8') as f:
            f.write(f"IMPRIMIR_CANCELADAS:{self.data_var.get()}")
        self.mostrar_toast("Gerando Relatório...", "info")

    def verificar_historico(self):
        try:
            with open(ARQUIVO_COMANDO, 'w', encoding='utf-8') as f:
                f.write("VERIFICAR_HISTORICO")
            self.mostrar_toast("Verificando historico...", "info")
        except Exception as e:
            self.mostrar_toast(f"Erro: {e}", "error")

    def atualizar_estoque_manual(self):
        try:
            with open(ARQUIVO_COMANDO, 'w', encoding='utf-8') as f:
                f.write("ATUALIZAR_ESTOQUE")
            self.mostrar_toast("Atualizando estoque...", "info")
        except Exception as e:
            self.mostrar_toast(f"Erro: {e}", "error")

    def enviar_print(self):
        t = self.ent_busca.get().strip()
        if t:
            with open(ARQUIVO_COMANDO, 'w', encoding='utf-8') as f:
                f.write(f"IMPRIMIR:{t}|{self.data_var.get()}")
            self.ent_busca.delete(0, "end")
            self.mostrar_toast(f"Imprimindo: {t}", "success")

    # ==================================================================================
    #  ALERTAS DE ATRASO (CONFIRMAÇÃO MANUAL)
    # ==================================================================================
    
    def carregar_alertas_atraso(self):
        """Carrega alertas pendentes e exibe na interface"""
        try:
            # Segurança: se a aba de vales ainda não foi inicializada, nada a fazer
            if not hasattr(self, 'fr_alertas_atraso'):
                return
            # Debounce: evita recarregar muito rápido
            now = time.time()
            if hasattr(self, '_last_alertas_load') and now - self._last_alertas_load < 2:
                return  # Ignora se carregou há menos de 2s
            self._last_alertas_load = now
            
            if not os.path.exists(ARQUIVO_ALERTAS):
                self.fr_alertas_atraso.grid_remove()
                return
            
            # Verifica se o arquivo mudou antes de recarregar
            try:
                mtime = os.path.getmtime(ARQUIVO_ALERTAS)
                if hasattr(self, '_alertas_mtime') and mtime == self._alertas_mtime:
                    return  # Arquivo não mudou
                self._alertas_mtime = mtime
            except OSError:
                return
            
            with open(ARQUIVO_ALERTAS, 'r', encoding='utf-8') as f:
                alertas = json.load(f)
            
            # Filtra apenas alertas de 9+ minutos (≥ 570 segundos)
            alertas_filtrados = []
            for alerta in alertas:
                tempo_total_s = int(alerta.get("tempo_minutos", 0)) * 60 + int(alerta.get("tempo_segundos", 0))
                if tempo_total_s >= 570:  # 9min30s ou mais
                    alertas_filtrados.append(alerta)
            
            if not alertas_filtrados:
                self.fr_alertas_atraso.grid_remove()
                return
            
            # Limpa alertas antigos da interface
            for widget in self.fr_alertas_atraso.winfo_children():
                widget.destroy()
            
            # Mostra o frame de alertas
            self.fr_alertas_atraso.grid()
            self.fr_alertas_atraso.configure(height=120 * min(len(alertas_filtrados), 3))  # Max 3 visíveis
            
            # Header
            ctk.CTkLabel(
                self.fr_alertas_atraso, 
                text="⚠️ Retirou com 9+ min - CONFIRME PARA ENVIAR",
                font=FONT_BOLD, 
                text_color=COR_DANGER
            ).pack(pady=(10, 5))
            
            # Card para cada alerta filtrado
            for alerta in alertas_filtrados:
                self._criar_card_alerta(alerta)
            
        except Exception as e:
            print(f"Erro ao carregar alertas: {e}")
    
    def _criar_card_alerta(self, alerta):
        """Cria um card para um alerta específico"""
        tempo_total_s = int(alerta.get("tempo_minutos", 0)) * 60 + int(alerta.get("tempo_segundos", 0))
        destaque_critico = tempo_total_s >= 570  # 9min30s
        cor_borda = COR_DANGER if destaque_critico else COR_DANGER
        cor_fundo = "#1a0f12" if destaque_critico else COR_BG_APP
        fr_alerta = ctk.CTkFrame(self.fr_alertas_atraso, fg_color=cor_fundo, corner_radius=8, border_width=1, border_color=cor_borda)
        fr_alerta.pack(fill="x", padx=10, pady=5)
        
        # Info principal
        fr_info = ctk.CTkFrame(fr_alerta, fg_color="transparent")
        fr_info.pack(side="left", fill="both", expand=True, padx=15, pady=10)
        
        # Pedido e cliente + selo
        fr_titulo = ctk.CTkFrame(fr_info, fg_color="transparent")
        fr_titulo.pack(anchor="w")
        ctk.CTkLabel(
            fr_titulo,
            text=f"#{alerta['numero']} - {alerta['cliente']}",
            font=("Roboto Bold", 14),
            text_color=COR_DANGER if destaque_critico else "white"
        ).pack(side="left")

        if destaque_critico:
            ctk.CTkLabel(
                fr_titulo,
                text="9,5+",
                font=("Roboto Bold", 11),
                text_color=COR_DANGER,
                fg_color="#2b1111",
                corner_radius=6,
                padx=6,
                pady=2
            ).pack(side="left", padx=(8, 0))
        
        # Tempo e motoboy
        # tipo_alerta = alerta.get("tipo", "atraso")  # unused variable removed
        tempo_txt = f"{alerta['tempo_minutos']}min {alerta['tempo_segundos']}s"
        info_txt = f"⏱️ Retirou com: {tempo_txt} | 🏍️ {alerta['motoboy']} | "
        info_txt += f"👥 {alerta['motoboys_livres']} livres / {alerta['motoboys_ocupados']} ocupados"
        
        ctk.CTkLabel(
            fr_info,
            text=info_txt,
            font=("Roboto", 11),
            text_color=COR_TEXT_SEC
        ).pack(anchor="w", pady=(2, 0))
        
        # Horários
        ctk.CTkLabel(
            fr_info,
            text=f"🕐 Pedido: {alerta['hora_aceito']} | Retirado: {alerta['timestamp']}",
            font=("Roboto", 10),
            text_color=COR_TEXT_SEC
        ).pack(anchor="w", pady=(2, 0))
        
        # Botões
        fr_btns = ctk.CTkFrame(fr_alerta, fg_color="transparent")
        fr_btns.pack(side="right", padx=10, pady=10)
        
        ctk.CTkButton(
            fr_btns,
            text="✅ ENVIAR",
            width=100,
            command=lambda a=alerta: self.enviar_alerta_atraso(a),
            fg_color=COR_SUCCESS,
            text_color="#003300",
            hover_color="#00A652"
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            fr_btns,
            text="❌ IGNORAR",
            width=100,
            command=lambda a=alerta: self.descartar_alerta(a),
            fg_color="transparent",
            border_width=1,
            border_color=COR_DANGER,
            text_color=COR_DANGER,
            hover_color="#2b1111"
        ).pack(side="left", padx=5)
    
    def enviar_alerta_atraso(self, alerta):
        """Confirma e envia o alerta para o WhatsApp via robô"""
        try:
            # Verifica se a menção está ativa
            mencao_ativa = self.config_data.get("whatsapp_mencao_ativa", False)
            mencao_txt = "@+55 49 9172-7951 " if mencao_ativa else ""
            
            # Monta mensagem formatada
            msg = (
                f"{mencao_txt}⚠️ Pedido: {alerta['numero']}\n"
                f"👤 *{alerta['cliente']}*\n"
                f"🏍️ Motoboy: {alerta['motoboy']}\n"
                f"⏱️ Retirou com: {alerta['tempo_minutos']}min {alerta['tempo_segundos']}s\n"
                f"🕐 Aceito: {alerta['hora_aceito']} | Saida: {alerta['timestamp']}"
            )
            
            # Envia comando para o robô enviar no WhatsApp
            with open(ARQUIVO_COMANDO, 'w', encoding='utf-8') as f:
                f.write(f"ENVIAR_WHATSAPP:{msg}")
            
            # Remove da lista
            self._remover_alerta_do_arquivo(alerta)
            
            self.mostrar_toast(f"Alerta #{alerta['numero']} enviado!", "success")
            
            # Recarrega interface
            self.after(100, self.carregar_alertas_atraso)
            
        except Exception as e:
            self.mostrar_toast(f"Erro: {e}", "error")
    
    def descartar_alerta(self, alerta):
        """Descarta o alerta sem enviar"""
        try:
            self._remover_alerta_do_arquivo(alerta)
            self.mostrar_toast(f"Alerta #{alerta['numero']} descartado", "info")
            self.after(100, self.carregar_alertas_atraso)
        except Exception as e:
            self.mostrar_toast(f"Erro: {e}", "error")
    
    def _remover_alerta_do_arquivo(self, alerta_remover):
        """Remove um alerta específico do arquivo JSON"""
        try:
            alertas = []
            if os.path.exists(ARQUIVO_ALERTAS):
                with open(ARQUIVO_ALERTAS, 'r', encoding='utf-8') as f:
                    alertas = json.load(f)
            
            # Filtra removendo o alerta específico
            alertas = [a for a in alertas if a['numero'] != alerta_remover['numero']]
            
            # Salva de volta
            with open(ARQUIVO_ALERTAS, 'w', encoding='utf-8') as f:
                json.dump(alertas, f, indent=2, ensure_ascii=False)
                
        except Exception as e:
            print(f"Erro ao remover alerta: {e}")

# ==================================================================================
#  SEÇÃO 14: INICIALIZAÇÃO DO PAINEL (MAIN LOOP)
# ==================================================================================
# Responsável por: Ponto de entrada principal da aplicação.
# Instancia Class PainelUltra e inicia a event loop do Tkinter.
# ==================================================================================

if __name__ == "__main__":
    app = PainelUltra()
    app.mainloop()